"""
実績見積Excel差分比較モジュール (CAD→概算見積もり生成AI / 日本住建株式会社)

AIが生成した概算見積 (estimate_calculator.Estimate) と、実際に顧客へ提出された
見積Excel（「16-古居徹朗様・古居百恵様邸新築工事【提出見積】.xlsx.xlsx」形式）を
突合し、差分を明らかにする。目的は以下の2つ:
  1. 画面での差分表示（総額・工事種別・明細の3階層）
  2. 見積精度較正用データの蓄積（パース結果はJSONシリアライズ可能な
     プリミティブのみで構成し、クラウドDBへそのまま保存できる）

提出見積xlsxの構造（「明細書」シート A1:H722 で確認済み）:
  - ヘッダ行: 項目№/名称/摘要/見積数量/単位/見積単価/見積金額
  - サマリ部: 項目№=1〜24 の工事種別サマリ行（単位=式）に続き、
    「端数値引き」（項目№=25等）「消費税等」「総合計」行が並ぶ
  - 明細部: 「総合計」行の後に種別ごとのブロックが続く。3階層構造:
      * 種別ヘッダ行  … 項目№=純数値（例 '1'）・金額なし
      * レベル1明細行 … 項目№='N-M'（例 '1-1'）・金額あり ← 種別合計と一致
      * レベル2内訳行 … 項目№='N-M-K'（例 '1-1-1'）… レベル1の内訳。
        サブヘッダ行（'N-M'・金額なし）に続いて現れ、集計すると
        レベル1と二重計上になるためパース対象から除外する
      * 「合計」行・注記行（※〜）はスキップ
  - 22管理費はサマリ行のみで明細ブロックを持たない

依存: openpyxl + 標準ライブラリのみ
"""

from __future__ import annotations

import io
import re
import unicodedata
from difflib import SequenceMatcher
from typing import Any

import openpyxl

# --- 定数 ---------------------------------------------------------------

DETAIL_SHEET_NAME = "明細書"          # 明細を読むシート名
COVER_SHEET_NAME = "表紙"             # 物件名を読むシート名
HEADER_SEARCH_ROWS = 10               # ヘッダ行を探す最大行数

# 明細マッチングのしきい値（過剰マッチより未マッチのほうが良い＝保守的に設定）
ITEM_MATCH_RATIO = 0.80               # difflib類似度の下限
ITEM_CONTAIN_MIN_LEN = 4              # 部分一致を認める正規化後の最小文字数
ITEM_CONTAIN_MIN_COVER = 0.6          # 部分一致で短い側が長い側を覆う最低比率
                                      # （「大工手間」⊂「●33調整分（大工手間）」の
                                      #   ような別項目への過剰マッチを防ぐ）
CATEGORY_MATCH_RATIO = 0.85           # 工事種別名のfuzzyマッチ下限
                                      # （外部建具/内部建具の誤マッチ0.83を弾く）

# 項目№のパターン
_RE_CAT_NO = re.compile(r"^\d+$")             # 種別ヘッダ (例 '1')
_RE_ITEM_NO = re.compile(r"^\d+-\d+$")        # レベル1明細 (例 '1-1')
_RE_SUBITEM_NO = re.compile(r"^\d+-\d+-\d+$")  # レベル2内訳 (例 '1-1-1')


# --- 内部ヘルパー -------------------------------------------------------

def _normalize_name(text: Any) -> str:
    """名称マッチング用に文字列を正規化する。

    NFKC正規化（半角カナ→全角・全角英数→半角）→ 小文字化 →
    空白/記号を除去し、英数字・かな・カタカナ・漢字のみを残す。
    例: 'ﾍﾞﾀ基礎' → 'ベタ基礎', '板金・樋工事' → '板金樋工事'
    """
    if text is None:
        return ""
    s = unicodedata.normalize("NFKC", str(text)).lower()
    # 英数字・ひらがな・カタカナ（長音符含む）・CJK漢字のみ残す
    return re.sub(r"[^0-9a-z぀-ヿ㐀-鿿豈-﫿]", "", s)


def _to_int(value: Any) -> int:
    """セル値を int に変換する（None・文字列は0）。"""
    if isinstance(value, bool):
        return 0
    if isinstance(value, (int, float)):
        return int(round(value))
    return 0


def _to_float(value: Any) -> float:
    """セル値を float に変換する（None・文字列は0.0）。"""
    if isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    return 0.0


def _cell_str(value: Any) -> str:
    """セル値を文字列化して前後空白を除去する。"""
    if value is None:
        return ""
    return str(value).strip()


def _extract_property_name(wb: openpyxl.Workbook) -> str:
    """表紙シートから「工事名」ラベルの右隣セルを物件名として取得する。"""
    if COVER_SHEET_NAME not in wb.sheetnames:
        return ""
    try:
        ws = wb[COVER_SHEET_NAME]
        for row in ws.iter_rows():
            cells = [c.value for c in row]
            for idx, val in enumerate(cells):
                if _cell_str(val) == "工事名":
                    # 同一行の右側で最初に見つかった非空セルを採用
                    for nxt in cells[idx + 1:]:
                        name = _cell_str(nxt)
                        if name:
                            return name
    except Exception:  # noqa: BLE001 - 表紙は装飾が多いため失敗しても致命でない
        pass
    return ""


def _similarity(a: str, b: str) -> float:
    """正規化済み文字列同士の類似度スコアを返す。

    完全一致=1.0 / 一定条件を満たす部分一致=0.90 / それ以外はdifflib比。
    部分一致は「短い側が一定長以上」かつ「長い側の6割以上を占める」場合のみ
    認める（短い名称が別項目の名称の一部に飲み込まれる過剰マッチを防ぐ）。
    """
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    shorter, longer = (a, b) if len(a) <= len(b) else (b, a)
    if (len(shorter) >= ITEM_CONTAIN_MIN_LEN
            and len(shorter) / len(longer) >= ITEM_CONTAIN_MIN_COVER
            and shorter in longer):
        return 0.90
    return SequenceMatcher(None, a, b).ratio()


# --- 公開 API: 実績見積のパース ------------------------------------------

def parse_actual_estimate(file_bytes: bytes) -> dict:
    """提出見積xlsxのバイト列をパースし、種別・明細を構造化して返す。

    Args:
        file_bytes: 提出見積Excelファイル (.xlsx) のバイト列。

    Returns:
        dict: JSONシリアライズ可能なプリミティブのみで構成された辞書。
            {
                'property_name': str,        # 表紙の工事名（取れなければ空文字）
                'total_excl_tax': int,       # 税抜合計（端数値引き込み・提出値）
                'total_incl_tax': int,       # 税込合計（総合計）
                'tax': int,                  # 消費税等
                'rounding_adjustment': int,  # 端数値引き（例 -206）
                'categories': [
                    {'no': int, 'name': str, 'total': int,
                     'items': [{'name': str, 'summary': str, 'quantity': float,
                                'unit': str, 'price': int, 'amount': int}]}
                ],
                'warnings': [str],           # 整合チェックで検出した注意事項
            }

    Raises:
        ValueError: 「明細書」シートが存在しない、またはヘッダ行が見つからない場合。
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if DETAIL_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"シート「{DETAIL_SHEET_NAME}」が見つかりません: {wb.sheetnames}")
    ws = wb[DETAIL_SHEET_NAME]

    # --- ヘッダ行（項目№/名称/...）を探す ---
    header_row = 0
    for r in range(1, HEADER_SEARCH_ROWS + 1):
        label = _normalize_name(ws.cell(row=r, column=1).value)
        if label.startswith("項目"):
            header_row = r
            break
    if header_row == 0:
        raise ValueError("「明細書」シートに項目№ヘッダ行が見つかりません")

    warnings: list[str] = []
    summary: dict[int, dict[str, Any]] = {}   # {種別no: {'name','total'}}
    items_by_cat: dict[int, list[dict]] = {}  # {種別no: [明細dict]}
    rounding_adjustment = 0
    tax = 0
    total_incl_tax = 0
    in_detail = False  # False=サマリ部 / True=明細部

    for r in range(header_row + 1, ws.max_row + 1):
        no_raw = _cell_str(ws.cell(row=r, column=1).value)
        name = _cell_str(ws.cell(row=r, column=2).value)
        amount_raw = ws.cell(row=r, column=7).value
        has_amount = isinstance(amount_raw, (int, float)) and not isinstance(amount_raw, bool)

        if not no_raw and not name:
            continue  # 空行

        # 項目№なし行: 消費税等・総合計・注記など
        if not no_raw:
            if "消費税" in name:
                tax = _to_int(amount_raw)
            elif name == "総合計":
                total_incl_tax = _to_int(amount_raw)
                in_detail = True  # 総合計以降が明細部
            # 「合計」行・「※〜」注記行はスキップ
            continue

        if _RE_CAT_NO.fullmatch(no_raw):
            if has_amount:
                # サマリ行（金額あり）。「端数値引き」等の調整行は種別から分離する
                if "値引" in name or "端数" in name:
                    rounding_adjustment += _to_int(amount_raw)
                else:
                    summary[int(no_raw)] = {"name": name, "total": _to_int(amount_raw)}
            else:
                # 明細部の種別ヘッダ行（金額なし）
                in_detail = True
            continue

        if in_detail and _RE_ITEM_NO.fullmatch(no_raw):
            if not has_amount:
                continue  # レベル2内訳のサブヘッダ行（'N-M'・金額なし）
            cat_no = int(no_raw.split("-")[0])
            items_by_cat.setdefault(cat_no, []).append({
                "name": name,
                "summary": _cell_str(ws.cell(row=r, column=3).value),
                "quantity": _to_float(ws.cell(row=r, column=4).value),
                "unit": _cell_str(ws.cell(row=r, column=5).value),
                "price": _to_int(ws.cell(row=r, column=6).value),
                "amount": _to_int(amount_raw),
            })
            continue

        # レベル2内訳行（'N-M-K'）は二重計上になるため集計しない
        if _RE_SUBITEM_NO.fullmatch(no_raw):
            continue

    # --- 種別リストを構築（サマリ部を正とし、明細を紐付ける）---
    categories: list[dict] = []
    for cat_no in sorted(set(summary) | set(items_by_cat)):
        info = summary.get(cat_no, {})
        items = items_by_cat.get(cat_no, [])
        total = _to_int(info.get("total", 0)) if info else sum(i["amount"] for i in items)
        name = info.get("name", "") or (f"種別{cat_no}")
        if not info:
            warnings.append(f"種別{cat_no}「{name}」はサマリ部に存在しません（明細のみ）")
        # 整合チェック: レベル1明細の合計 = サマリ金額
        if items and info:
            item_sum = sum(i["amount"] for i in items)
            if item_sum != total:
                warnings.append(
                    f"種別{cat_no}「{name}」の明細合計{item_sum:,}円が"
                    f"サマリ金額{total:,}円と一致しません"
                )
        categories.append({"no": cat_no, "name": name, "total": total, "items": items})

    # --- 総額の確定と整合チェック ---
    category_sum = sum(c["total"] for c in categories)
    total_excl_tax = category_sum + rounding_adjustment  # 提出値（端数値引き込み）
    if total_incl_tax and tax and total_incl_tax - tax != total_excl_tax:
        warnings.append(
            f"税込{total_incl_tax:,}円-消費税{tax:,}円={total_incl_tax - tax:,}円が"
            f"税抜合計{total_excl_tax:,}円と一致しません"
        )
    if not total_incl_tax:
        total_incl_tax = total_excl_tax + tax

    return {
        "property_name": _extract_property_name(wb),
        "total_excl_tax": total_excl_tax,
        "total_incl_tax": total_incl_tax,
        "tax": tax,
        "rounding_adjustment": rounding_adjustment,
        "categories": categories,
        "warnings": warnings,
    }


# --- 内部ヘルパー: マッチング ---------------------------------------------

def _match_categories(ai_cats: list, actual_cats: list) -> list[tuple]:
    """AI種別と実績種別を名称ベースで対応付ける。

    Returns:
        list[tuple]: (ai_category | None, actual_category(dict) | None) のリスト。
            片方にしか無い種別は相手側を None として含める。
    """
    pairs: list[tuple] = []
    used_actual: set[int] = set()

    # 1) 正規化名の完全一致
    actual_by_norm = {}
    for idx, ac in enumerate(actual_cats):
        actual_by_norm.setdefault(_normalize_name(ac["name"]), idx)
    unmatched_ai = []
    for cat in ai_cats:
        idx = actual_by_norm.get(_normalize_name(cat.name))
        if idx is not None and idx not in used_actual:
            pairs.append((cat, actual_cats[idx]))
            used_actual.add(idx)
        else:
            unmatched_ai.append(cat)

    # 2) 残りをfuzzyマッチ（高類似のみ・スコア降順に貪欲確定）
    candidates = []
    for cat in unmatched_ai:
        for idx, ac in enumerate(actual_cats):
            if idx in used_actual:
                continue
            score = _similarity(_normalize_name(cat.name), _normalize_name(ac["name"]))
            if score >= CATEGORY_MATCH_RATIO:
                candidates.append((score, id(cat), cat, idx))
    matched_ai_ids: set[int] = set()
    for score, cat_id, cat, idx in sorted(candidates, key=lambda t: -t[0]):
        if cat_id in matched_ai_ids or idx in used_actual:
            continue
        pairs.append((cat, actual_cats[idx]))
        matched_ai_ids.add(cat_id)
        used_actual.add(idx)

    # 3) どちらにも対応が付かなかった種別
    for cat in unmatched_ai:
        if id(cat) not in matched_ai_ids:
            pairs.append((cat, None))
    for idx, ac in enumerate(actual_cats):
        if idx not in used_actual:
            pairs.append((None, ac))

    return pairs


def _match_items(ai_items: list, actual_items: list) -> dict:
    """種別内の明細同士を名称類似度で対応付ける。

    完全一致 → 部分一致（正規化後4文字以上）→ difflib類似度0.80以上の順で
    スコアリングし、スコア降順の貪欲法で1対1に確定する。
    しきい値未満は無理にマッチさせない（保守的）。
    """
    scored = []
    for i, ai in enumerate(ai_items):
        norm_ai = _normalize_name(ai.name)
        for j, ac in enumerate(actual_items):
            score = _similarity(norm_ai, _normalize_name(ac["name"]))
            if score >= ITEM_MATCH_RATIO:
                scored.append((score, i, j))

    matched = []
    used_ai: set[int] = set()
    used_actual: set[int] = set()
    for score, i, j in sorted(scored, key=lambda t: -t[0]):
        if i in used_ai or j in used_actual:
            continue
        ai, ac = ai_items[i], actual_items[j]
        matched.append({
            "ai_name": ai.name,
            "actual_name": ac["name"],
            "ai_amount": int(ai.estimate_amount),
            "actual_amount": int(ac["amount"]),
            "diff": int(ai.estimate_amount) - int(ac["amount"]),
        })
        used_ai.add(i)
        used_actual.add(j)

    ai_only = [{"name": ai.name, "amount": int(ai.estimate_amount)}
               for i, ai in enumerate(ai_items) if i not in used_ai]
    actual_only = [{"name": ac["name"], "amount": int(ac["amount"])}
                   for j, ac in enumerate(actual_items) if j not in used_actual]
    return {"matched": matched, "ai_only": ai_only, "actual_only": actual_only}


# --- 公開 API: 差分計算 ---------------------------------------------------

def diff_estimates(estimate, actual: dict) -> dict:
    """AI概算見積と実績見積パース結果の差分を計算する。

    Args:
        estimate: estimate_calculator.Estimate オブジェクト（AI概算見積）。
        actual: parse_actual_estimate() の戻り値（実績見積）。

    Returns:
        dict: JSONシリアライズ可能な差分構造。
            {
                'total_ai': int, 'total_actual': int,
                'diff': int,                  # total_ai - total_actual
                'rate': float,                # diff / total_actual
                'category_diffs': [           # 名称マッピング済み・片側のみも含む
                    {'no', 'name', 'ai_total', 'actual_total', 'diff', 'rate'}
                ],
                'item_diffs': {
                    種別名: {'matched': [{'ai_name','actual_name',
                                          'ai_amount','actual_amount','diff'}],
                             'ai_only': [{'name','amount'}],
                             'actual_only': [{'name','amount'}]}
                },
            }
    """
    total_ai = int(estimate.total_estimate_excl_tax)
    total_actual = int(actual["total_excl_tax"])
    diff = total_ai - total_actual
    rate = diff / total_actual if total_actual else 0.0

    category_diffs: list[dict] = []
    item_diffs: dict[str, dict] = {}

    for ai_cat, actual_cat in _match_categories(list(estimate.categories),
                                                list(actual["categories"])):
        ai_total = int(ai_cat.estimate_total) if ai_cat is not None else 0
        actual_total = int(actual_cat["total"]) if actual_cat is not None else 0
        cat_no = ai_cat.no if ai_cat is not None else actual_cat["no"]
        cat_name = ai_cat.name if ai_cat is not None else actual_cat["name"]
        cat_diff = ai_total - actual_total
        category_diffs.append({
            "no": int(cat_no),
            "name": cat_name,
            "ai_total": ai_total,
            "actual_total": actual_total,
            "diff": cat_diff,
            "rate": cat_diff / actual_total if actual_total else 0.0,
        })
        # 明細差分（どちらか一方しか無い種別も ai_only / actual_only で表現）
        ai_items = list(ai_cat.items) if ai_cat is not None else []
        actual_items = list(actual_cat["items"]) if actual_cat is not None else []
        if ai_items or actual_items:
            item_diffs[cat_name] = _match_items(ai_items, actual_items)

    category_diffs.sort(key=lambda d: d["no"])

    return {
        "total_ai": total_ai,
        "total_actual": total_actual,
        "diff": diff,
        "rate": rate,
        "category_diffs": category_diffs,
        "item_diffs": item_diffs,
    }


# --- 公開 API: 日本語サマリ -----------------------------------------------

def diff_summary_text(diff: dict) -> str:
    """差分結果からフィードバック送信用の日本語サマリテキストを生成する。

    総額差・率、差の大きい工事種別トップ5、代表的な明細差（金額差の大きい
    マッチ明細＋片側にしか無い主要明細）をプレーンテキストで整形する。
    """
    lines: list[str] = []
    lines.append("【実績見積との差分サマリ】")
    lines.append("■ 総額（税抜）")
    lines.append(f"  AI概算: {diff['total_ai']:,}円 / 実績: {diff['total_actual']:,}円")
    lines.append(f"  差額: {diff['diff']:+,}円（{diff['rate']:+.1%}）")

    # 差の大きい工事種別トップ5
    top_cats = sorted(diff["category_diffs"], key=lambda d: -abs(d["diff"]))[:5]
    lines.append("■ 差の大きい工事種別 TOP5")
    for rank, cat in enumerate(top_cats, start=1):
        rate_str = f" / {cat['rate']:+.1%}" if cat["actual_total"] else ""
        lines.append(
            f"  {rank}. {cat['name']}: AI {cat['ai_total']:,}円 / "
            f"実績 {cat['actual_total']:,}円（差 {cat['diff']:+,}円{rate_str}）"
        )

    # 代表的な明細差: マッチ済みで金額差の大きいものトップ5
    matched_all = []
    ai_only_all = []
    actual_only_all = []
    for cat_name, detail in diff["item_diffs"].items():
        for m in detail["matched"]:
            if m["diff"] != 0:
                matched_all.append((cat_name, m))
        for it in detail["ai_only"]:
            ai_only_all.append((cat_name, it))
        for it in detail["actual_only"]:
            actual_only_all.append((cat_name, it))

    lines.append("■ 代表的な明細差")
    for cat_name, m in sorted(matched_all, key=lambda t: -abs(t[1]["diff"]))[:5]:
        lines.append(
            f"  [{cat_name}] {m['ai_name']}: AI {m['ai_amount']:,}円 / "
            f"実績 {m['actual_amount']:,}円（差 {m['diff']:+,}円）"
        )
    for cat_name, it in sorted(ai_only_all, key=lambda t: -abs(t[1]["amount"]))[:3]:
        lines.append(f"  （AIのみ）[{cat_name}] {it['name']}: {it['amount']:,}円")
    for cat_name, it in sorted(actual_only_all, key=lambda t: -abs(t[1]["amount"]))[:3]:
        lines.append(f"  （実績のみ）[{cat_name}] {it['name']}: {it['amount']:,}円")
    if not matched_all and not ai_only_all and not actual_only_all:
        lines.append("  （明細レベルの差分はありません）")

    return "\n".join(lines)


# --- CLI エントリポイント -----------------------------------------------

if __name__ == "__main__":
    # 古居様邸の実データで動作確認する（開発用）
    import json
    from pathlib import Path

    base = Path(__file__).resolve().parent
    xlsx_path = base / "16-古居徹朗様・古居百恵様邸新築工事【提出見積】.xlsx.xlsx"

    print("=== excel_diff.py 動作確認 ===")
    actual = parse_actual_estimate(xlsx_path.read_bytes())
    print(f"物件名: {actual['property_name']}")
    print(f"税抜合計: {actual['total_excl_tax']:,}円 / 税込合計: {actual['total_incl_tax']:,}円")
    print(f"種別数: {len(actual['categories'])} / "
          f"明細数: {sum(len(c['items']) for c in actual['categories'])}")
    for w in actual["warnings"]:
        print(f"  警告: {w}")

    from cad_parser import load_cad_file
    from estimate_calculator import calculate_estimate, load_unit_prices

    cad = load_cad_file(str(base / "16-古居徹朗様・古居百恵様邸新築工事【数量データ】.TXT"))
    est = calculate_estimate(cad, load_unit_prices(str(base / "tankamaster_updated.csv")))
    result = diff_estimates(est, actual)
    print(f"\n差分率: {result['rate']:+.2%}")
    print()
    print(diff_summary_text(result))
    # JSONシリアライズ可能であることの確認
    json.dumps(actual, ensure_ascii=False)
    json.dumps(result, ensure_ascii=False)
    print("\nJSONシリアライズ: OK")
