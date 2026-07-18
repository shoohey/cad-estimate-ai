"""日本住建 見積AI システム - デモアプリケーション"""
import streamlit as st
import pandas as pd
import os
import io
import csv
import math
import copy
import json
import base64
import hashlib
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

from cad_parser import load_cad_from_bytes, load_cad_file, CADData
from estimate_calculator import calculate_estimate, load_unit_prices, Estimate, EstimateItem
from andpad_converter import (convert_to_andpad, ANDPADBudget,
                              ANDPADAllocation, vendor_sort_key)
from feedback_handler import submit_feedback as send_feedback_to_company
import cloud_store

# ページ設定
st.set_page_config(
    page_title="日本住建 見積AI",
    page_icon="🏠",
    layout="wide",
    initial_sidebar_state="expanded",
)

# CSS - 高可読性デザイン
st.markdown("""
<style>
    /* ===== ライトテーマ強制 ===== */
    :root {
        color-scheme: light !important;
    }

    /* ===== 全体ベース ===== */
    .stApp {
        background-color: #f8f9fb !important;
        font-family: "Hiragino Kaku Gothic ProN", "Noto Sans JP", "Yu Gothic", "Meiryo", sans-serif;
        color: #23272f !important;
    }

    /* ===== サイドバー ===== */
    section[data-testid="stSidebar"] {
        background-color: #1a2744 !important;
        border-right: none;
    }
    section[data-testid="stSidebar"] p,
    section[data-testid="stSidebar"] label,
    section[data-testid="stSidebar"] h1,
    section[data-testid="stSidebar"] h2,
    section[data-testid="stSidebar"] h3,
    section[data-testid="stSidebar"] h4,
    section[data-testid="stSidebar"] h5,
    section[data-testid="stSidebar"] span,
    section[data-testid="stSidebar"] div,
    section[data-testid="stSidebar"] li,
    section[data-testid="stSidebar"] summary {
        color: #e2e8f0 !important;
    }
    section[data-testid="stSidebar"] .stMarkdown h5 {
        color: #f1f5f9 !important;
        font-weight: 700 !important;
        font-size: 0.9rem !important;
        letter-spacing: 0.04em;
        border-bottom: 1px solid #334155;
        padding-bottom: 0.3rem;
    }
    section[data-testid="stSidebar"] .stCheckbox label p,
    section[data-testid="stSidebar"] .stCheckbox label span {
        color: #cbd5e1 !important;
    }
    section[data-testid="stSidebar"] input,
    section[data-testid="stSidebar"] textarea,
    section[data-testid="stSidebar"] select {
        color: #23272f !important;
        background-color: #ffffff !important;
    }
    section[data-testid="stSidebar"] [data-baseweb="input"],
    section[data-testid="stSidebar"] [data-baseweb="textarea"],
    section[data-testid="stSidebar"] [data-baseweb="select"] > div {
        background-color: #ffffff !important;
        border-color: #475569 !important;
    }
    section[data-testid="stSidebar"] hr {
        border-color: #334155 !important;
    }
    /* サイドバーのヘルプ？アイコン */
    section[data-testid="stSidebar"] .stTooltipHoverTarget button,
    section[data-testid="stSidebar"] [data-testid="stTooltipHoverTarget"] button {
        color: #ffffff !important;
        background-color: transparent !important;
        border: 1px solid rgba(255,255,255,0.3) !important;
    }
    section[data-testid="stSidebar"] .stTooltipHoverTarget button svg,
    section[data-testid="stSidebar"] [data-testid="stTooltipHoverTarget"] button svg {
        stroke: #ffffff !important;
    }
    /* --- サイドバーボタン --- */
    section[data-testid="stSidebar"] button {
        color: #ffffff !important;
        border: 1px solid rgba(255,255,255,0.25) !important;
        background-color: rgba(255,255,255,0.12) !important;
        font-weight: 600 !important;
        border-radius: 8px !important;
        padding: 0.5rem 1rem !important;
    }
    section[data-testid="stSidebar"] button:hover {
        background-color: rgba(255,255,255,0.2) !important;
        border-color: rgba(255,255,255,0.4) !important;
    }
    section[data-testid="stSidebar"] button[kind="primary"] {
        background-color: #3b82f6 !important;
        border-color: #3b82f6 !important;
    }
    section[data-testid="stSidebar"] button[kind="primary"]:hover {
        background-color: #2563eb !important;
        border-color: #2563eb !important;
    }

    /* ===== フォント ===== */
    .stApp p, .stApp label, .stApp li,
    .stApp input, .stApp textarea, .stApp select,
    .stApp td, .stApp th,
    .stApp h1, .stApp h2, .stApp h3, .stApp h4, .stApp h5, .stApp h6,
    .stApp [data-testid="stMarkdownContainer"],
    .stApp [data-testid="stText"] {
        font-family: "Hiragino Kaku Gothic ProN", "Noto Sans JP", "Yu Gothic", "Meiryo", sans-serif;
    }

    /* ===== メインエリア文字色 ===== */
    section.stMain p, section.stMain label, section.stMain li,
    section.stMain td, section.stMain th,
    section.stMain input, section.stMain textarea, section.stMain select,
    section.stMain summary,
    section.stMain h1, section.stMain h2, section.stMain h3,
    section.stMain h4, section.stMain h5, section.stMain h6,
    section.stMain [data-testid="stMarkdownContainer"],
    section.stMain [data-testid="stText"] {
        color: #23272f !important;
    }

    /* リンク色 */
    .stApp a:not([class*="icon"]):not([data-baseweb]),
    .stApp a:visited:not([class*="icon"]):not([data-baseweb]) {
        color: #2563eb !important;
        text-decoration: underline !important;
    }
    .stApp a:hover:not([class*="icon"]):not([data-baseweb]) {
        color: #1d4ed8 !important;
    }

    /* ===== ヘッダー ===== */
    .main-header {
        background: linear-gradient(135deg, #1e3a5f 0%, #2c4a7c 50%, #1a3055 100%);
        padding: 1.5rem 2rem;
        border-radius: 12px;
        margin-bottom: 1.5rem;
        border-bottom: 3px solid #c4a265;
        box-shadow: 0 4px 12px rgba(26,48,85,0.2);
    }
    section.stMain .main-header h1,
    section.stMain .main-header h2,
    section.stMain .main-header h3,
    section.stMain .main-header p,
    section.stMain .main-header span,
    section.stMain .main-header div,
    section.stMain .main-header label,
    section.stMain .main-header a,
    .stApp .main-header h1,
    .stApp .main-header p,
    .stApp .main-header span,
    .stApp .main-header div {
        color: #ffffff !important;
    }
    section.stMain .main-header h1,
    .stApp .main-header h1 {
        margin: 0; font-size: 1.5rem; font-weight: 700;
        color: #ffffff !important; letter-spacing: 0.08em;
        text-shadow: 0 1px 3px rgba(0,0,0,0.25);
    }
    section.stMain .main-header p,
    .stApp .main-header p {
        margin: 0.4rem 0 0 0; color: #ffffff !important;
        font-size: 0.85rem; letter-spacing: 0.04em;
        opacity: 0.9;
    }

    /* ===== 案件選択バー ===== */
    .project-bar {
        background: #ffffff; border: 1px solid #e2e8f0; border-radius: 10px;
        padding: 0.6rem 1.2rem; margin-bottom: 1rem;
        box-shadow: 0 1px 3px rgba(0,0,0,0.04);
        display: flex; align-items: center; gap: 0.8rem;
    }
    .project-badge {
        display: inline-block; background: #e0f2fe; color: #0c4a6e !important;
        font-size: 0.78rem; font-weight: 700; padding: 3px 12px; border-radius: 6px;
    }
    .project-badge-active { background: #2c5282; color: #ffffff !important; }

    /* ===== メトリクスカード ===== */
    div[data-testid="stMetric"] {
        background: #ffffff; border: 1px solid #e2e8f0;
        border-radius: 10px; padding: 16px 18px;
        box-shadow: 0 1px 4px rgba(0,0,0,0.04);
    }
    div[data-testid="stMetric"] label {
        color: #64748b !important; font-size: 0.82rem !important;
        font-weight: 600 !important;
    }
    div[data-testid="stMetric"] [data-testid="stMetricValue"] {
        color: #0f172a !important; font-size: 1.35rem !important; font-weight: 700 !important;
    }
    div[data-testid="stMetric"] [data-testid="stMetricDelta"] {
        font-size: 0.8rem !important; color: #64748b !important;
    }

    /* ===== 確認事項・警告ボックス ===== */
    .warning-box {
        background: #fffbeb; border: 1px solid #fde68a; border-left: 4px solid #f59e0b;
        padding: 0.8rem 1.1rem; margin: 0.5rem 0; border-radius: 0 8px 8px 0;
        color: #92400e !important; font-size: 0.88rem; line-height: 1.65;
    }
    .warning-box p, .warning-box span, .warning-box div,
    .warning-box a, .warning-box br { color: #92400e !important; }

    /* ===== info-box ===== */
    .info-box {
        background: #eff6ff; border: 1px solid #bfdbfe; border-left: 4px solid #3b82f6;
        padding: 0.8rem 1.1rem; margin: 0.5rem 0; border-radius: 0 8px 8px 0;
        color: #1e3a5f !important; font-size: 0.88rem; line-height: 1.65;
    }
    .info-box p, .info-box span, .info-box div { color: #1e3a5f !important; }

    /* ===== タブ ===== */
    .stTabs [data-baseweb="tab-list"] {
        gap: 2px; border-bottom: 2px solid #e2e8f0; background: transparent;
    }
    .stTabs [data-baseweb="tab"] {
        padding: 10px 20px; border-radius: 8px 8px 0 0; font-size: 0.88rem;
        font-weight: 600; color: #64748b !important;
        border: 1px solid transparent; border-bottom: none; background: transparent;
    }
    .stTabs [data-baseweb="tab"]:hover {
        color: #1e3a5f !important; background: #f1f5f9;
    }
    .stTabs [data-baseweb="tab"][aria-selected="true"] {
        color: #1e3a5f !important; border-color: #e2e8f0;
        border-bottom: 2px solid #f8f9fb; background: #f8f9fb;
    }

    /* ===== セクション見出し ===== */
    .stMarkdown h4 {
        color: #1e3a5f !important; font-weight: 700 !important; font-size: 1.08rem !important;
        padding-bottom: 0.4rem; border-bottom: 2px solid #e2e8f0;
        margin-bottom: 1rem !important; margin-top: 0.5rem !important;
    }
    section.stMain .stMarkdown h5 {
        color: #334155 !important; font-weight: 600 !important; font-size: 0.95rem !important;
    }

    /* ===== テーブル / DataFrame ===== */
    .stDataFrame { border: 1px solid #e2e8f0; border-radius: 8px; overflow: hidden; }
    .stApp [data-testid="stDataFrame"] [role="columnheader"],
    .stApp [data-testid="stDataFrame"] [role="columnheader"] span {
        background-color: #f1f5f9 !important; color: #1e3a5f !important; font-weight: 600 !important;
    }
    .stApp [data-testid="stDataFrame"] [role="gridcell"],
    .stApp [data-testid="stDataFrame"] [role="gridcell"] span {
        color: #23272f !important; background-color: #ffffff !important;
    }

    /* ==========================================================
       ボタン — section.stMain で正しくスコープ
       ========================================================== */

    /* --- メインエリア: デフォルトボタン --- */
    section.stMain button[kind="secondary"],
    section.stMain button[kind="primary"],
    section.stMain button[kind="minimal"],
    section.stMain [data-testid="stBaseButton-secondary"] button,
    section.stMain [data-testid="stBaseButton-primary"] button,
    section.stMain .stButton button,
    section.stMain .stDownloadButton button,
    section.stMain .stFormSubmitButton button,
    section.stMain [data-testid="stFormSubmitButton"] button {
        border-radius: 8px !important;
        font-weight: 600 !important;
        font-size: 0.86rem !important;
        padding: 0.5rem 1.2rem !important;
        cursor: pointer !important;
        line-height: 1.5 !important;
    }

    /* --- 通常ボタン（secondary = Streamlitのデフォルト） --- */
    section.stMain button[kind="secondary"] {
        background-color: #ffffff !important;
        color: #1e3a5f !important;
        border: 1.5px solid #cbd5e1 !important;
        box-shadow: 0 1px 2px rgba(0,0,0,0.04) !important;
    }
    section.stMain button[kind="secondary"]:hover {
        background-color: #eff6ff !important;
        border-color: #93c5fd !important;
        color: #1e40af !important;
    }

    /* --- Primaryボタン（保存・適用・送信・開く） --- */
    section.stMain button[kind="primary"] {
        background-color: #2563eb !important;
        color: #ffffff !important;
        border: 1.5px solid #2563eb !important;
        box-shadow: 0 1px 3px rgba(37,99,235,0.18) !important;
    }
    section.stMain button[kind="primary"]:hover {
        background-color: #1d4ed8 !important;
        border-color: #1d4ed8 !important;
        color: #ffffff !important;
        box-shadow: 0 2px 8px rgba(37,99,235,0.28) !important;
    }

    /* --- ボタン内テキスト色の強制上書き --- */
    section.stMain button[kind="secondary"] p,
    section.stMain button[kind="secondary"] span,
    section.stMain button[kind="secondary"] div {
        color: #1e3a5f !important;
    }
    section.stMain button[kind="primary"] p,
    section.stMain button[kind="primary"] span,
    section.stMain button[kind="primary"] div {
        color: #ffffff !important;
    }

    /* --- ダウンロードボタン（secondary = 通常DL） --- */
    section.stMain .stDownloadButton button[kind="secondary"],
    section.stMain [data-testid="stDownloadButton"] button[kind="secondary"] {
        background-color: #f0fdfa !important;
        color: #0f766e !important;
        border: 1.5px solid #99f6e4 !important;
    }
    section.stMain .stDownloadButton button[kind="secondary"]:hover,
    section.stMain [data-testid="stDownloadButton"] button[kind="secondary"]:hover {
        background-color: #ccfbf1 !important;
        border-color: #5eead4 !important;
        color: #115e59 !important;
    }
    section.stMain .stDownloadButton button[kind="secondary"] p,
    section.stMain .stDownloadButton button[kind="secondary"] span,
    section.stMain [data-testid="stDownloadButton"] button[kind="secondary"] p,
    section.stMain [data-testid="stDownloadButton"] button[kind="secondary"] span {
        color: #0f766e !important;
    }

    /* --- ダウンロードボタン（primary = 重要DL） --- */
    section.stMain .stDownloadButton button[kind="primary"],
    section.stMain [data-testid="stDownloadButton"] button[kind="primary"] {
        background-color: #0d9488 !important;
        color: #ffffff !important;
        border: 1.5px solid #0d9488 !important;
    }
    section.stMain .stDownloadButton button[kind="primary"]:hover,
    section.stMain [data-testid="stDownloadButton"] button[kind="primary"]:hover {
        background-color: #0f766e !important;
        border-color: #0f766e !important;
        color: #ffffff !important;
    }
    section.stMain .stDownloadButton button[kind="primary"] p,
    section.stMain .stDownloadButton button[kind="primary"] span,
    section.stMain [data-testid="stDownloadButton"] button[kind="primary"] p,
    section.stMain [data-testid="stDownloadButton"] button[kind="primary"] span {
        color: #ffffff !important;
    }

    /* ===== ファイルアップローダー ===== */
    section.stMain [data-testid="stFileUploader"] {
        border: 2px dashed #94a3b8; border-radius: 10px;
        padding: 1rem; background: #f1f5f9;
    }
    section.stMain [data-testid="stFileUploaderDropzone"] {
        background: #ffffff !important;
        border: 1px dashed #94a3b8 !important; border-radius: 8px;
    }

    /* ===== divider ===== */
    hr { border-color: #e2e8f0 !important; }

    /* ===== 入力フィールド ===== */
    section.stMain input, section.stMain textarea, section.stMain select {
        background-color: #ffffff !important; color: #23272f !important;
    }
    section.stMain [data-baseweb="input"],
    section.stMain [data-baseweb="textarea"] {
        background-color: #ffffff !important; border-color: #cbd5e1 !important;
    }
    section.stMain [data-baseweb="select"] > div {
        background-color: #ffffff !important; border-color: #cbd5e1 !important;
    }

    /* ===== number_input +/- ===== */
    section.stMain [data-testid="stNumberInput"] button,
    .stApp [data-testid="stNumberInput"] button {
        color: #475569 !important;
        background-color: #f1f5f9 !important;
        border: 1px solid #cbd5e1 !important;
        padding: 0.25rem 0.5rem !important;
        box-shadow: none !important;
    }
    section.stMain [data-testid="stNumberInput"] button:hover {
        background-color: #e2e8f0 !important; color: #1e3a5f !important;
    }

    /* ===== data_editor ===== */
    .stApp [data-testid="stDataEditor"] {
        border: 1px solid #e2e8f0; border-radius: 8px; overflow: hidden;
    }
    .stApp [data-testid="stDataEditor"] [role="columnheader"],
    .stApp [data-testid="stDataEditor"] [role="columnheader"] span {
        color: #1e3a5f !important; background-color: #f1f5f9 !important;
    }
    .stApp [data-testid="stDataEditor"] [role="gridcell"],
    .stApp [data-testid="stDataEditor"] [role="gridcell"] span,
    .stApp [data-testid="stDataEditor"] [role="gridcell"] input {
        color: #23272f !important;
    }

    /* ===== alertボックス ===== */
    .stApp [data-testid="stAlert"] p,
    .stApp [data-testid="stAlert"] span,
    .stApp [data-testid="stAlert"] div { color: #23272f !important; }

    /* ===== expander ===== */
    .stApp [data-testid="stExpander"] {
        border: 1px solid #e2e8f0 !important; border-radius: 10px;
        background: #ffffff; margin-bottom: 0.4rem;
    }
    .stApp [data-testid="stExpander"] summary p {
        color: #23272f !important; font-weight: 600 !important; font-size: 0.9rem !important;
    }
    .stApp [data-testid="stExpander"] [data-testid="stMarkdownContainer"] p {
        font-weight: 400 !important;
    }

    /* ===== json viewer ===== */
    .stApp [data-testid="stJson"] { color: #23272f !important; }

    /* ===== フォーム ===== */
    .stApp [data-testid="stForm"] {
        background: #f8fafc; border: 1px solid #e2e8f0;
        border-radius: 10px; padding: 1.2rem;
    }

    /* ===== file uploader text ===== */
    section.stMain [data-testid="stFileUploader"] p,
    section.stMain [data-testid="stFileUploader"] span,
    section.stMain [data-testid="stFileUploader"] label,
    section.stMain [data-testid="stFileUploader"] div,
    section.stMain [data-testid="stFileUploader"] small,
    section.stMain [data-testid="stFileUploaderDropzone"] p,
    section.stMain [data-testid="stFileUploaderDropzone"] span {
        color: #334155 !important;
    }
    section.stMain [data-testid="stFileUploader"] > label p {
        color: #1e293b !important; font-weight: 600 !important; font-size: 0.95rem !important;
    }

    /* ===== radio / checkbox / select / slider ===== */
    .stApp [role="radiogroup"] label p,
    .stApp [role="radiogroup"] label span { color: #23272f !important; }
    .stApp [data-testid="stCheckbox"] label p,
    .stApp [data-testid="stCheckbox"] label span { color: #23272f !important; }
    .stApp [data-baseweb="select"] span,
    .stApp [data-baseweb="select"] div { color: #23272f !important; }
    .stApp [data-testid="stSlider"] p,
    .stApp [data-testid="stSlider"] div { color: #23272f !important; }

    /* ===== feedback card ===== */
    .feedback-card {
        background: #ffffff;
        border: 1px solid #e2e8f0;
        border-radius: 10px;
        padding: 1rem 1.2rem;
        margin-bottom: 0.5rem;
        box-shadow: 0 1px 3px rgba(0,0,0,0.03);
    }
    .feedback-card .fb-meta {
        font-size: 0.78rem;
        color: #64748b !important;
        margin-bottom: 0.3rem;
    }
    .feedback-card .fb-body {
        color: #23272f !important;
        font-size: 0.9rem;
        line-height: 1.6;
    }
    .feedback-card .fb-category {
        display: inline-block;
        background: #dbeafe;
        color: #1e3a5f !important;
        font-size: 0.75rem;
        font-weight: 600;
        padding: 2px 10px;
        border-radius: 4px;
        margin-right: 0.4rem;
    }

    /* ===== ウェルカム画面のステップカード ===== */
    .step-card {
        background: #ffffff;
        border: 1px solid #e2e8f0;
        border-radius: 12px;
        padding: 1.4rem 1.6rem;
        height: 100%;
        box-shadow: 0 2px 6px rgba(0,0,0,0.04);
    }
    .step-card .step-num {
        display: inline-block;
        background: #2c5282;
        color: #ffffff !important;
        font-weight: 700;
        font-size: 0.78rem;
        padding: 3px 12px;
        border-radius: 6px;
        margin-bottom: 0.6rem;
        letter-spacing: 0.06em;
    }
    .step-card h3 {
        color: #1e3a5f !important;
        font-size: 1.02rem;
        font-weight: 700;
        margin: 0.4rem 0 0.5rem 0;
    }
    .step-card p {
        color: #475569 !important;
        font-size: 0.88rem;
        line-height: 1.65;
        margin: 0;
    }
</style>
""", unsafe_allow_html=True)


def format_currency(amount: int) -> str:
    """金額をカンマ区切りフォーマット"""
    if amount < 0:
        return f"-¥{abs(amount):,}"
    return f"¥{amount:,}"


def format_number(num: float) -> str:
    """数値のフォーマット"""
    if num == int(num):
        return f"{int(num):,}"
    return f"{num:,.2f}"


def create_excel_estimate(estimate: Estimate) -> bytes:
    """Excel見積書を生成"""
    wb = Workbook()

    # ===== 表紙シート =====
    ws_cover = wb.active
    ws_cover.title = "表紙"
    ws_cover.sheet_properties.pageSetUpPr = None

    # スタイル定義
    title_font = Font(name='游ゴシック', size=18, bold=True)
    header_font = Font(name='游ゴシック', size=12, bold=True)
    normal_font = Font(name='游ゴシック', size=11)
    amount_font = Font(name='游ゴシック', size=16, bold=True, color='1a365d')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    header_fill = PatternFill(start_color='1a365d', end_color='1a365d', fill_type='solid')
    header_font_white = Font(name='游ゴシック', size=11, bold=True, color='FFFFFF')

    # 列幅設定
    ws_cover.column_dimensions['A'].width = 5
    ws_cover.column_dimensions['B'].width = 20
    ws_cover.column_dimensions['C'].width = 30
    ws_cover.column_dimensions['D'].width = 20
    ws_cover.column_dimensions['E'].width = 20

    # 表紙内容
    ws_cover.merge_cells('B2:D2')
    ws_cover['B2'] = '概 算 見 積 書'
    ws_cover['B2'].font = title_font
    ws_cover['B2'].alignment = Alignment(horizontal='center')

    ws_cover['B4'] = '見積番号'
    ws_cover['C4'] = estimate.estimate_no or ''
    ws_cover['B5'] = '提出日'
    ws_cover['C5'] = datetime.now().strftime('%Y年%m月%d日')
    ws_cover['B7'] = '施主名'
    ws_cover['C7'] = f'{estimate.owner_name} 様' if estimate.owner_name else ''
    ws_cover['B8'] = '工事名'
    ws_cover['C8'] = estimate.property_name
    ws_cover['B9'] = '工事場所'
    ws_cover['C9'] = estimate.location or ''

    ws_cover.merge_cells('B11:D11')
    ws_cover['B11'] = '見積金額（税込）'
    ws_cover['B11'].font = header_font
    ws_cover.merge_cells('B12:D12')
    ws_cover['B12'] = f'¥{estimate.total_estimate_incl_tax:,}'
    ws_cover['B12'].font = amount_font

    ws_cover['B14'] = '税抜金額'
    ws_cover['C14'] = estimate.total_estimate_excl_tax
    ws_cover['C14'].number_format = '#,##0'
    ws_cover['B15'] = '消費税(10%)'
    ws_cover['C15'] = estimate.tax
    ws_cover['C15'].number_format = '#,##0'
    ws_cover['B17'] = '延床面積'
    ws_cover['C17'] = f'{estimate.total_floor_area_m2:.2f}㎡ ({estimate.total_floor_area_tsubo:.2f}坪)'

    for row in range(4, 18):
        ws_cover[f'B{row}'].font = normal_font
        ws_cover[f'C{row}'].font = normal_font

    # ===== 明細シート =====
    ws_detail = wb.create_sheet("明細書")

    # 列幅
    col_widths = {'A': 5, 'B': 6, 'C': 30, 'D': 25, 'E': 10, 'F': 6,
                  'G': 12, 'H': 15, 'I': 12, 'J': 15}
    for col, width in col_widths.items():
        ws_detail.column_dimensions[col].width = width

    # ヘッダー
    headers = ['No', '項番', '名称', '摘要', '数量', '単位',
               '見積単価', '見積金額', '発注単価', '発注金額']
    for col_idx, header in enumerate(headers, 1):
        cell = ws_detail.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    row_num = 2
    seq = 1
    for category in estimate.categories:
        # 工事種別ヘッダー
        cat_fill = PatternFill(start_color='e2e8f0', end_color='e2e8f0', fill_type='solid')
        ws_detail.cell(row=row_num, column=1, value=category.no).font = Font(bold=True)
        ws_detail.merge_cells(start_row=row_num, start_column=2,
                             end_row=row_num, end_column=3)
        ws_detail.cell(row=row_num, column=2, value=category.name).font = Font(bold=True)
        ws_detail.cell(row=row_num, column=8, value=category.estimate_total)
        ws_detail.cell(row=row_num, column=8).number_format = '#,##0'
        ws_detail.cell(row=row_num, column=8).font = Font(bold=True)
        ws_detail.cell(row=row_num, column=10, value=category.order_total)
        ws_detail.cell(row=row_num, column=10).number_format = '#,##0'
        ws_detail.cell(row=row_num, column=10).font = Font(bold=True)
        for col in range(1, 11):
            ws_detail.cell(row=row_num, column=col).fill = cat_fill
            ws_detail.cell(row=row_num, column=col).border = thin_border
        row_num += 1

        # 明細
        for item in category.items:
            ws_detail.cell(row=row_num, column=2, value=seq)
            ws_detail.cell(row=row_num, column=3, value=item.name)
            ws_detail.cell(row=row_num, column=4, value=item.summary)
            ws_detail.cell(row=row_num, column=5, value=item.quantity)
            ws_detail.cell(row=row_num, column=5).number_format = '#,##0.00'
            ws_detail.cell(row=row_num, column=6, value=item.unit)
            ws_detail.cell(row=row_num, column=7, value=item.estimate_price)
            ws_detail.cell(row=row_num, column=7).number_format = '#,##0'
            ws_detail.cell(row=row_num, column=8, value=item.estimate_amount)
            ws_detail.cell(row=row_num, column=8).number_format = '#,##0'
            ws_detail.cell(row=row_num, column=9, value=item.order_price)
            ws_detail.cell(row=row_num, column=9).number_format = '#,##0'
            ws_detail.cell(row=row_num, column=10, value=item.order_amount)
            ws_detail.cell(row=row_num, column=10).number_format = '#,##0'
            for col in range(1, 11):
                ws_detail.cell(row=row_num, column=col).border = thin_border
                ws_detail.cell(row=row_num, column=col).font = normal_font
            row_num += 1
            seq += 1

    # 合計行
    row_num += 1
    ws_detail.merge_cells(start_row=row_num, start_column=2,
                         end_row=row_num, end_column=6)
    ws_detail.cell(row=row_num, column=2, value='税抜合計').font = Font(bold=True, size=12)
    ws_detail.cell(row=row_num, column=8, value=estimate.total_estimate_excl_tax)
    ws_detail.cell(row=row_num, column=8).number_format = '#,##0'
    ws_detail.cell(row=row_num, column=8).font = Font(bold=True, size=12)
    ws_detail.cell(row=row_num, column=10, value=estimate.total_order_excl_tax)
    ws_detail.cell(row=row_num, column=10).number_format = '#,##0'
    ws_detail.cell(row=row_num, column=10).font = Font(bold=True, size=12)

    row_num += 1
    ws_detail.cell(row=row_num, column=2, value='消費税(10%)').font = Font(bold=True)
    ws_detail.cell(row=row_num, column=8, value=estimate.tax)
    ws_detail.cell(row=row_num, column=8).number_format = '#,##0'

    row_num += 1
    ws_detail.cell(row=row_num, column=2, value='税込合計').font = Font(bold=True, size=14, color='1a365d')
    ws_detail.cell(row=row_num, column=8, value=estimate.total_estimate_incl_tax)
    ws_detail.cell(row=row_num, column=8).number_format = '#,##0'
    ws_detail.cell(row=row_num, column=8).font = Font(bold=True, size=14, color='1a365d')

    # ===== 粗利表シート =====
    ws_profit = wb.create_sheet("粗利表")
    profit_headers = ['No', '工事種別', '見積金額', '発注金額', '粗利額', '粗利率']
    for col_idx, header in enumerate(profit_headers, 1):
        cell = ws_profit.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    ws_profit.column_dimensions['A'].width = 5
    ws_profit.column_dimensions['B'].width = 25
    ws_profit.column_dimensions['C'].width = 15
    ws_profit.column_dimensions['D'].width = 15
    ws_profit.column_dimensions['E'].width = 15
    ws_profit.column_dimensions['F'].width = 10

    row_num = 2
    for cat in estimate.categories:
        ws_profit.cell(row=row_num, column=1, value=cat.no)
        ws_profit.cell(row=row_num, column=2, value=cat.name)
        ws_profit.cell(row=row_num, column=3, value=cat.estimate_total)
        ws_profit.cell(row=row_num, column=3).number_format = '#,##0'
        ws_profit.cell(row=row_num, column=4, value=cat.order_total)
        ws_profit.cell(row=row_num, column=4).number_format = '#,##0'
        ws_profit.cell(row=row_num, column=5, value=cat.profit)
        ws_profit.cell(row=row_num, column=5).number_format = '#,##0'
        ws_profit.cell(row=row_num, column=6, value=cat.profit_rate)
        ws_profit.cell(row=row_num, column=6).number_format = '0.0%'
        for col in range(1, 7):
            ws_profit.cell(row=row_num, column=col).border = thin_border
        row_num += 1

    # 合計行
    ws_profit.cell(row=row_num, column=2, value='合計').font = Font(bold=True)
    ws_profit.cell(row=row_num, column=3, value=estimate.total_estimate_excl_tax)
    ws_profit.cell(row=row_num, column=3).number_format = '#,##0'
    ws_profit.cell(row=row_num, column=3).font = Font(bold=True)
    ws_profit.cell(row=row_num, column=4, value=estimate.total_order_excl_tax)
    ws_profit.cell(row=row_num, column=4).number_format = '#,##0'
    ws_profit.cell(row=row_num, column=4).font = Font(bold=True)
    ws_profit.cell(row=row_num, column=5, value=estimate.total_profit)
    ws_profit.cell(row=row_num, column=5).number_format = '#,##0'
    ws_profit.cell(row=row_num, column=5).font = Font(bold=True)
    ws_profit.cell(row=row_num, column=6, value=estimate.total_profit_rate)
    ws_profit.cell(row=row_num, column=6).number_format = '0.0%'
    ws_profit.cell(row=row_num, column=6).font = Font(bold=True)

    # バイト出力
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_andpad_excel(budget: ANDPADBudget) -> bytes:
    """ANDPAD用Excel工務予算を生成"""
    wb = Workbook()
    ws = wb.active
    ws.title = "工務予算（材工分離）"

    header_fill = PatternFill(start_color='2d5a87', end_color='2d5a87', fill_type='solid')
    header_font = Font(name='游ゴシック', size=11, bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    headers = ['工事種別', '細目工種（発注先）', '名称', '摘要',
               '数量', '単位', '材料費', '施工費', '発注金額']
    col_widths = [18, 25, 30, 25, 10, 6, 14, 14, 14]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    row_num = 2
    for item in budget.items:
        multi = len(item.allocations) > 1
        for alloc in item.allocations:
            summary = item.summary
            if multi and alloc.note:
                summary = f"{summary}　※{alloc.note}" if summary else f"※{alloc.note}"
            ws.cell(row=row_num, column=1, value=item.work_category)
            ws.cell(row=row_num, column=2, value=alloc.vendor)
            ws.cell(row=row_num, column=3, value=item.item_name)
            ws.cell(row=row_num, column=4, value=summary)
            ws.cell(row=row_num, column=5, value=item.quantity)
            ws.cell(row=row_num, column=5).number_format = '#,##0.00'
            ws.cell(row=row_num, column=6, value=item.unit)
            ws.cell(row=row_num, column=7, value=alloc.amount if alloc.kind == '材' else 0)
            ws.cell(row=row_num, column=7).number_format = '#,##0'
            ws.cell(row=row_num, column=8, value=alloc.amount if alloc.kind == '工' else 0)
            ws.cell(row=row_num, column=8).number_format = '#,##0'
            ws.cell(row=row_num, column=9, value=alloc.amount)
            ws.cell(row=row_num, column=9).number_format = '#,##0'
            for col in range(1, 10):
                ws.cell(row=row_num, column=col).border = thin_border
            row_num += 1

    # 合計
    row_num += 1
    ws.cell(row=row_num, column=2, value='合計').font = Font(bold=True)
    ws.cell(row=row_num, column=7, value=budget.total_material)
    ws.cell(row=row_num, column=7).number_format = '#,##0'
    ws.cell(row=row_num, column=7).font = Font(bold=True)
    ws.cell(row=row_num, column=8, value=budget.total_labor)
    ws.cell(row=row_num, column=8).number_format = '#,##0'
    ws.cell(row=row_num, column=8).font = Font(bold=True)
    ws.cell(row=row_num, column=9, value=budget.grand_total)
    ws.cell(row=row_num, column=9).number_format = '#,##0'
    ws.cell(row=row_num, column=9).font = Font(bold=True)

    # 発注先別集計シート
    ws2 = wb.create_sheet("発注先別集計")
    headers2 = ['発注先', '材料費', '施工費', '合計']
    for col_idx, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 14
    ws2.column_dimensions['D'].width = 14

    by_vendor = budget.by_vendor()
    row_num = 2
    for vendor, data in by_vendor.items():
        ws2.cell(row=row_num, column=1, value=vendor)
        ws2.cell(row=row_num, column=2, value=data['material'])
        ws2.cell(row=row_num, column=2).number_format = '#,##0'
        ws2.cell(row=row_num, column=3, value=data['labor'])
        ws2.cell(row=row_num, column=3).number_format = '#,##0'
        ws2.cell(row=row_num, column=4, value=data['total'])
        ws2.cell(row=row_num, column=4).number_format = '#,##0'
        for col in range(1, 5):
            ws2.cell(row=row_num, column=col).border = thin_border
        row_num += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_summary_csv(estimate: Estimate) -> bytes:
    """工事別内訳CSVを生成"""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['No', '工事種別', '見積金額', '発注金額', '粗利額', '粗利率'])
    for cat in estimate.categories:
        pr = f"{cat.profit_rate * 100:.1f}%" if cat.estimate_total > 0 else "-"
        writer.writerow([cat.no, cat.name, cat.estimate_total, cat.order_total, cat.profit, pr])
    writer.writerow(['', '税抜合計', estimate.total_estimate_excl_tax, estimate.total_order_excl_tax,
                     estimate.total_profit, f"{estimate.total_profit_rate * 100:.1f}%"])
    writer.writerow(['', '消費税(10%)', estimate.tax, '', '', ''])
    writer.writerow(['', '税込合計', estimate.total_estimate_incl_tax, '', '', ''])
    return output.getvalue().encode('utf-8-sig')


def create_detail_csv(estimate: Estimate) -> bytes:
    """見積明細CSVを生成"""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['工事種別', 'カテゴリ', '名称', '摘要', '数量', '単位',
                     '見積単価', '見積金額', '発注単価', '発注金額'])
    for cat in estimate.categories:
        for item in cat.items:
            writer.writerow([
                f"{cat.no}. {cat.name}", item.category or '-', item.name, item.summary,
                f"{item.quantity:.2f}", item.unit,
                item.estimate_price, item.estimate_amount,
                item.order_price, item.order_amount,
            ])
    return output.getvalue().encode('utf-8-sig')


def create_vendor_csv(budget: ANDPADBudget) -> bytes:
    """発注先別集計CSVを生成（明細付き・予実管理表の並び順）"""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['発注先（細目工種）', '材料費', '施工費', '合計', '明細数'])
    by_vendor = budget.by_vendor()
    for vendor, data in by_vendor.items():
        writer.writerow([vendor, data['material'], data['labor'], data['total'],
                         len(data['details'])])
    writer.writerow(['合計', budget.total_material, budget.total_labor,
                     budget.grand_total, ''])
    # 発注先別の明細（ANDPAD発注時のコピー＆ペースト用）
    writer.writerow([])
    writer.writerow(['【発注先別明細】'])
    for vendor, data in by_vendor.items():
        writer.writerow([])
        writer.writerow([vendor, '', '', f"合計 {data['total']}", ''])
        writer.writerow(['工事種別', '名称', '摘要', '数量', '単位', '材料費', '施工費', '金額'])
        for item, alloc in data['details']:
            summary = item.summary
            if len(item.allocations) > 1 and alloc.note:
                summary = f"{summary}　※{alloc.note}" if summary else f"※{alloc.note}"
            writer.writerow([
                item.work_category, item.item_name, summary,
                f"{item.quantity:.2f}", item.unit,
                alloc.amount if alloc.kind == '材' else 0,
                alloc.amount if alloc.kind == '工' else 0,
                alloc.amount,
            ])
    return output.getvalue().encode('utf-8-sig')


def create_andpad_csv(budget: ANDPADBudget) -> bytes:
    """ANDPADインポート用CSVを生成（材工分離後の配分単位で1行）"""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['工事種別', '細目工種（発注先）', '名称', '摘要',
                     '数量', '単位', '材料費', '施工費', '発注金額'])
    for item in budget.items:
        multi = len(item.allocations) > 1
        for alloc in item.allocations:
            summary = item.summary
            if multi and alloc.note:
                summary = f"{summary}　※{alloc.note}" if summary else f"※{alloc.note}"
            writer.writerow([
                item.work_category,
                alloc.vendor,
                item.item_name,
                summary,
                f"{item.quantity:.2f}",
                item.unit,
                alloc.amount if alloc.kind == '材' else 0,
                alloc.amount if alloc.kind == '工' else 0,
                alloc.amount,
            ])
    return output.getvalue().encode('utf-8-sig')


def apply_modifications(estimate: Estimate, mods: dict) -> Estimate:
    """修正データを見積に適用"""
    overrides = mods.get('overrides', {})
    deleted = mods.get('deleted', set())
    new_items = mods.get('new_items', [])

    for cat in estimate.categories:
        if cat.no == 22:
            continue
        items_to_keep = []
        for i, item in enumerate(cat.items):
            key = f"{cat.no}_{i}"
            if key in deleted:
                continue
            if key in overrides:
                ov = overrides[key]
                item.quantity = ov['quantity']
                item.estimate_price = ov['estimate_price']
                item.order_price = ov['order_price']
                if 'name' in ov:
                    item.name = ov['name']
                if 'summary' in ov:
                    item.summary = ov['summary']
                item.estimate_amount = int(item.quantity * item.estimate_price)
                item.order_amount = int(item.quantity * item.order_price)
            items_to_keep.append(item)
        cat.items = items_to_keep

        for new in new_items:
            if new['cat_no'] == cat.no:
                ni = EstimateItem(
                    parent_no=cat.no, child_no=0,
                    name=new['name'], summary=new.get('summary', ''),
                    quantity=new['quantity'], unit=new['unit'],
                    estimate_price=new['estimate_price'],
                    order_price=new['order_price'],
                    estimate_amount=int(new['quantity'] * new['estimate_price']),
                    order_amount=int(new['quantity'] * new['order_price']),
                )
                cat.items.append(ni)

    # 管理費再計算
    for cat in estimate.categories:
        if cat.no == 22:
            mgmt_est = estimate.management_fee_estimate
            mgmt_ord = estimate.management_fee_order
            cat.items = [EstimateItem(
                parent_no=22, child_no=1,
                name="管理費", summary="小計×3%/2%", unit="式",
                quantity=1,
                estimate_price=mgmt_est, order_price=mgmt_ord,
                estimate_amount=mgmt_est, order_amount=mgmt_ord,
            )]

    return estimate


def apply_profit_rate_adjustment(estimate: Estimate, target_rate: float) -> Estimate:
    """目標粗利率に合わせて見積単価を比例調整する。
    元の見積金額の比率を維持しつつ、全体の粗利率が目標値になるよう
    一律の倍率をかける。
    倍率 m = total_order / (total_estimate_current * (1 - target_rate))
    """
    if target_rate <= 0 or target_rate >= 1:
        return estimate

    # 管理費（cat22）を除いた現在の見積・発注合計
    current_est = sum(c.estimate_total for c in estimate.categories if c.no != 22)
    current_ord = sum(c.order_total for c in estimate.categories if c.no != 22)

    if current_est == 0 or current_ord == 0:
        return estimate

    # 目標粗利率を達成するための倍率を算出
    # target_rate = (m * current_est - current_ord) / (m * current_est)
    # → m = current_ord / (current_est * (1 - target_rate))
    multiplier = current_ord / (current_est * (1 - target_rate))

    for cat in estimate.categories:
        if cat.no == 22:
            continue
        for item in cat.items:
            item.estimate_price = int(item.estimate_price * multiplier)
            item.estimate_amount = int(item.quantity * item.estimate_price)

    # 管理費再計算
    for cat in estimate.categories:
        if cat.no == 22:
            mgmt_est = estimate.management_fee_estimate
            mgmt_ord = estimate.management_fee_order
            cat.items = [EstimateItem(
                parent_no=22, child_no=1,
                name="管理費", summary="小計×3%/2%", unit="式",
                quantity=1,
                estimate_price=mgmt_est, order_price=mgmt_ord,
                estimate_amount=mgmt_est, order_amount=mgmt_ord,
            )]

    return estimate


def create_client_excel(estimate: Estimate) -> bytes:
    """お客様提出用Excel見積書（発注金額・粗利情報なし）"""
    wb = Workbook()

    # スタイル定義
    title_font = Font(name='游ゴシック', size=18, bold=True)
    header_font = Font(name='游ゴシック', size=12, bold=True)
    normal_font = Font(name='游ゴシック', size=11)
    amount_font = Font(name='游ゴシック', size=16, bold=True, color='1a365d')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    header_fill = PatternFill(start_color='1a365d', end_color='1a365d', fill_type='solid')
    header_font_white = Font(name='游ゴシック', size=11, bold=True, color='FFFFFF')

    # ===== 表紙シート =====
    ws_cover = wb.active
    ws_cover.title = "表紙"
    ws_cover.column_dimensions['A'].width = 5
    ws_cover.column_dimensions['B'].width = 20
    ws_cover.column_dimensions['C'].width = 30
    ws_cover.column_dimensions['D'].width = 20

    ws_cover.merge_cells('B2:D2')
    ws_cover['B2'] = '概 算 見 積 書'
    ws_cover['B2'].font = title_font
    ws_cover['B2'].alignment = Alignment(horizontal='center')

    ws_cover['B4'] = '見積番号'
    ws_cover['C4'] = estimate.estimate_no or ''
    ws_cover['B5'] = '提出日'
    ws_cover['C5'] = datetime.now().strftime('%Y年%m月%d日')
    ws_cover['B7'] = '施主名'
    ws_cover['C7'] = f'{estimate.owner_name} 様' if estimate.owner_name else ''
    ws_cover['B8'] = '工事名'
    ws_cover['C8'] = estimate.property_name
    ws_cover['B9'] = '工事場所'
    ws_cover['C9'] = estimate.location or ''

    ws_cover.merge_cells('B11:D11')
    ws_cover['B11'] = '見積金額（税込）'
    ws_cover['B11'].font = header_font
    ws_cover.merge_cells('B12:D12')
    ws_cover['B12'] = f'¥{estimate.total_estimate_incl_tax:,}'
    ws_cover['B12'].font = amount_font

    ws_cover['B14'] = '税抜金額'
    ws_cover['C14'] = estimate.total_estimate_excl_tax
    ws_cover['C14'].number_format = '#,##0'
    ws_cover['B15'] = '消費税(10%)'
    ws_cover['C15'] = estimate.tax
    ws_cover['C15'].number_format = '#,##0'
    ws_cover['B17'] = '延床面積'
    ws_cover['C17'] = f'{estimate.total_floor_area_m2:.2f}㎡ ({estimate.total_floor_area_tsubo:.2f}坪)'

    for row in range(4, 18):
        ws_cover[f'B{row}'].font = normal_font
        ws_cover[f'C{row}'].font = normal_font

    # ===== 明細シート（見積側のみ） =====
    ws_detail = wb.create_sheet("明細書")
    col_widths = {'A': 5, 'B': 6, 'C': 30, 'D': 25, 'E': 10, 'F': 6, 'G': 12, 'H': 15}
    for col, width in col_widths.items():
        ws_detail.column_dimensions[col].width = width

    headers = ['No', '項番', '名称', '摘要', '数量', '単位', '単価', '金額']
    for col_idx, header in enumerate(headers, 1):
        cell = ws_detail.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    row_num = 2
    seq = 1
    for category in estimate.categories:
        cat_fill = PatternFill(start_color='e2e8f0', end_color='e2e8f0', fill_type='solid')
        ws_detail.cell(row=row_num, column=1, value=category.no).font = Font(bold=True)
        ws_detail.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3)
        ws_detail.cell(row=row_num, column=2, value=category.name).font = Font(bold=True)
        ws_detail.cell(row=row_num, column=8, value=category.estimate_total)
        ws_detail.cell(row=row_num, column=8).number_format = '#,##0'
        ws_detail.cell(row=row_num, column=8).font = Font(bold=True)
        for col in range(1, 9):
            ws_detail.cell(row=row_num, column=col).fill = cat_fill
            ws_detail.cell(row=row_num, column=col).border = thin_border
        row_num += 1

        for item in category.items:
            ws_detail.cell(row=row_num, column=2, value=seq)
            ws_detail.cell(row=row_num, column=3, value=item.name)
            ws_detail.cell(row=row_num, column=4, value=item.summary)
            ws_detail.cell(row=row_num, column=5, value=item.quantity)
            ws_detail.cell(row=row_num, column=5).number_format = '#,##0.00'
            ws_detail.cell(row=row_num, column=6, value=item.unit)
            ws_detail.cell(row=row_num, column=7, value=item.estimate_price)
            ws_detail.cell(row=row_num, column=7).number_format = '#,##0'
            ws_detail.cell(row=row_num, column=8, value=item.estimate_amount)
            ws_detail.cell(row=row_num, column=8).number_format = '#,##0'
            for col in range(1, 9):
                ws_detail.cell(row=row_num, column=col).border = thin_border
                ws_detail.cell(row=row_num, column=col).font = normal_font
            row_num += 1
            seq += 1

    # 合計行
    row_num += 1
    ws_detail.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=6)
    ws_detail.cell(row=row_num, column=2, value='税抜合計').font = Font(bold=True, size=12)
    ws_detail.cell(row=row_num, column=8, value=estimate.total_estimate_excl_tax)
    ws_detail.cell(row=row_num, column=8).number_format = '#,##0'
    ws_detail.cell(row=row_num, column=8).font = Font(bold=True, size=12)

    row_num += 1
    ws_detail.cell(row=row_num, column=2, value='消費税(10%)').font = Font(bold=True)
    ws_detail.cell(row=row_num, column=8, value=estimate.tax)
    ws_detail.cell(row=row_num, column=8).number_format = '#,##0'

    row_num += 1
    ws_detail.cell(row=row_num, column=2, value='税込合計').font = Font(bold=True, size=14, color='1a365d')
    ws_detail.cell(row=row_num, column=8, value=estimate.total_estimate_incl_tax)
    ws_detail.cell(row=row_num, column=8).number_format = '#,##0'
    ws_detail.cell(row=row_num, column=8).font = Font(bold=True, size=14, color='1a365d')

    # ===== 工事別集計シート =====
    ws_summary = wb.create_sheet("工事別集計")
    sum_headers = ['No', '工事種別', '金額']
    for col_idx, header in enumerate(sum_headers, 1):
        cell = ws_summary.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
    ws_summary.column_dimensions['A'].width = 5
    ws_summary.column_dimensions['B'].width = 25
    ws_summary.column_dimensions['C'].width = 15

    row_num = 2
    for cat in estimate.categories:
        ws_summary.cell(row=row_num, column=1, value=cat.no)
        ws_summary.cell(row=row_num, column=2, value=cat.name)
        ws_summary.cell(row=row_num, column=3, value=cat.estimate_total)
        ws_summary.cell(row=row_num, column=3).number_format = '#,##0'
        for col in range(1, 4):
            ws_summary.cell(row=row_num, column=col).border = thin_border
        row_num += 1

    ws_summary.cell(row=row_num, column=2, value='合計').font = Font(bold=True)
    ws_summary.cell(row=row_num, column=3, value=estimate.total_estimate_excl_tax)
    ws_summary.cell(row=row_num, column=3).number_format = '#,##0'
    ws_summary.cell(row=row_num, column=3).font = Font(bold=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


PROJECTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "projects")
os.makedirs(PROJECTS_DIR, exist_ok=True)


def _warn_cloud_error(exc: Exception):
    """クラウド接続エラー時: 以降このセッションはローカルのみで動作させ、バナーで通知する。

    サーキットブレーカー: 一度エラーを検知したら _cloud_ok() が False になり、
    以降のクラウド読み書きを全てスキップする（タイムアウトの連鎖でUIが
    ブロックするのを防ぎ、古いローカルindexでクラウドを上書きする事故も防ぐ）。
    バナーは main() 冒頭で毎rerun表示される。
    """
    st.session_state['_cloud_degraded'] = True
    st.session_state['_cloud_error_msg'] = str(exc)
    st.warning(f"クラウド保存に接続できませんでした。変更はアプリ再起動時に失われる可能性があります。({exc})")


def _cloud_ok() -> bool:
    """クラウド永続化が使用可能か（構成済み かつ このセッションで障害未検知）"""
    return cloud_store.enabled() and not st.session_state.get('_cloud_degraded')


def _load_project_list_local() -> dict:
    index_path = os.path.join(PROJECTS_DIR, "index.json")
    if os.path.exists(index_path):
        with open(index_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def _seed_cloud_from_local(local_index: dict):
    """クラウド初回起動時、リポジトリ同梱のローカルデータを初期投入する"""
    try:
        for pid in local_index:
            dpath = os.path.join(PROJECTS_DIR, f"{pid}_data.json")
            if os.path.exists(dpath):
                with open(dpath, 'r', encoding='utf-8') as f:
                    cloud_store.put(f"proj:{pid}:data", json.load(f))
            cpath = os.path.join(PROJECTS_DIR, f"{pid}.txt")
            if os.path.exists(cpath):
                with open(cpath, 'rb') as f:
                    cloud_store.put(f"proj:{pid}:cad",
                                    {"b64": base64.b64encode(f.read()).decode('ascii')})
            fpath = os.path.join(PROJECTS_DIR, f"{pid}_feedbacks.json")
            if os.path.exists(fpath):
                with open(fpath, 'r', encoding='utf-8') as f:
                    cloud_store.put(f"proj:{pid}:feedbacks", json.load(f))
        # indexは最後に書く（途中で失敗した場合は次回起動時に再シードされる）
        cloud_store.put("index", local_index)
    except cloud_store.CloudStoreError as exc:
        _warn_cloud_error(exc)


def load_project_list() -> dict:
    """保存済み案件一覧を読み込む（クラウド優先・ローカルフォールバック）"""
    local = _load_project_list_local()
    if _cloud_ok():
        try:
            cloud = cloud_store.get("index")
            if cloud is None:
                _seed_cloud_from_local(local)
                return local
            # jsonbはキー順を保持しないため、作成日時順（従来の登録順）に並べ直す
            return dict(sorted(cloud.items(), key=lambda kv: (kv[1] or {}).get('created', '')))
        except cloud_store.CloudStoreError as exc:
            _warn_cloud_error(exc)
    return local


def save_project_list(projects: dict):
    """案件一覧を保存（ローカル＋クラウド）"""
    index_path = os.path.join(PROJECTS_DIR, "index.json")
    with open(index_path, 'w', encoding='utf-8') as f:
        json.dump(projects, f, ensure_ascii=False, indent=2)
    if _cloud_ok():
        try:
            cloud_store.put("index", projects)
        except cloud_store.CloudStoreError as exc:
            _warn_cloud_error(exc)


def _convert_sets(obj):
    """再帰的にsetをJSON互換形式に変換"""
    if isinstance(obj, set):
        return {'__set__': list(obj)}
    if isinstance(obj, dict):
        return {k: _convert_sets(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_convert_sets(item) for item in obj]
    return obj


def save_project_data(proj_id: str, data: dict):
    """案件の修正データ等を保存（ローカル＋クラウド）"""
    path = os.path.join(PROJECTS_DIR, f"{proj_id}_data.json")
    save_data = _convert_sets(data)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(save_data, f, ensure_ascii=False, indent=2)
    if _cloud_ok():
        # 毎rerunの自動保存で内容が変わっていない時はクラウド送信をスキップする
        digest = hashlib.sha1(
            json.dumps(save_data, ensure_ascii=False, sort_keys=True).encode('utf-8')
        ).hexdigest()
        cache_key = f'_cloud_saved_{proj_id}'
        if st.session_state.get(cache_key) == digest:
            return
        try:
            cloud_store.put(f"proj:{proj_id}:data", save_data)
            st.session_state[cache_key] = digest
        except cloud_store.CloudStoreError as exc:
            _warn_cloud_error(exc)


def load_project_data(proj_id: str) -> dict:
    """案件の修正データ等を読み込む（クラウド優先・ローカルフォールバック）"""
    if _cloud_ok():
        try:
            val = cloud_store.get(f"proj:{proj_id}:data")
            if val is not None:
                return _restore_sets(val)
        except cloud_store.CloudStoreError as exc:
            _warn_cloud_error(exc)
    path = os.path.join(PROJECTS_DIR, f"{proj_id}_data.json")
    if os.path.exists(path):
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return _restore_sets(data)
    return {}


def _restore_sets(obj):
    """再帰的に__set__マーカーをsetに復元"""
    if isinstance(obj, dict):
        if '__set__' in obj and len(obj) == 1:
            return set(obj['__set__'])
        return {k: _restore_sets(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_restore_sets(item) for item in obj]
    return obj


def save_cad_file(proj_id: str, file_bytes: bytes):
    """CADファイルを保存（ローカル＋クラウド）"""
    path = os.path.join(PROJECTS_DIR, f"{proj_id}.txt")
    with open(path, 'wb') as f:
        f.write(file_bytes)
    if _cloud_ok():
        try:
            cloud_store.put(f"proj:{proj_id}:cad",
                            {"b64": base64.b64encode(file_bytes).decode('ascii')})
        except cloud_store.CloudStoreError as exc:
            _warn_cloud_error(exc)


def load_project_feedbacks(proj_id: str) -> list:
    """案件ごとのフィードバックを読み込む（クラウド優先・ローカルフォールバック）

    フィードバックタブは毎rerunで読み直すため、セッション内キャッシュで
    クラウドへのGET連発を抑える（保存・削除時にキャッシュを更新）。
    """
    cache = st.session_state.setdefault('_fb_cache', {})
    if proj_id in cache:
        return cache[proj_id]
    result = None
    if _cloud_ok():
        try:
            val = cloud_store.get(f"proj:{proj_id}:feedbacks")
            if val is not None:
                result = val
        except cloud_store.CloudStoreError as exc:
            _warn_cloud_error(exc)
    if result is None:
        path = os.path.join(PROJECTS_DIR, f"{proj_id}_feedbacks.json")
        if os.path.exists(path):
            with open(path, 'r', encoding='utf-8') as f:
                result = json.load(f)
        else:
            result = []
    cache[proj_id] = result
    return result


def save_project_feedbacks(proj_id: str, fbs: list):
    """案件ごとのフィードバックを保存（ローカル＋クラウド）"""
    path = os.path.join(PROJECTS_DIR, f"{proj_id}_feedbacks.json")
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(fbs, f, ensure_ascii=False, indent=2)
    st.session_state.setdefault('_fb_cache', {})[proj_id] = fbs
    if _cloud_ok():
        try:
            cloud_store.put(f"proj:{proj_id}:feedbacks", fbs)
        except cloud_store.CloudStoreError as exc:
            _warn_cloud_error(exc)


def load_all_feedbacks(projects: dict) -> list:
    """全案件のフィードバックを読み込み、案件IDを付与して返す"""
    all_fbs = []
    for pid in projects:
        for fb in load_project_feedbacks(pid):
            # キャッシュ上のdictを汚さないようコピーに表示用キーを付与する
            fb2 = dict(fb)
            fb2['_proj_id'] = pid
            fb2['_proj_name'] = projects[pid].get('name', pid)
            all_fbs.append(fb2)
    all_fbs.sort(key=lambda x: x.get('timestamp', ''), reverse=True)
    return all_fbs


def migrate_global_feedbacks(projects: dict, base_dir: str):
    """旧グローバルfeedbacks.jsonを案件別に移行"""
    fb_file = os.path.join(base_dir, "feedbacks.json")
    if not os.path.exists(fb_file):
        return
    with open(fb_file, 'r', encoding='utf-8') as f:
        old_fbs = json.load(f)
    if not old_fbs:
        os.rename(fb_file, fb_file + ".bak")
        return
    # property名で案件を特定して振り分け
    proj_by_prop = {}
    for pid, pinfo in projects.items():
        proj_by_prop[pinfo.get('name', '')] = pid
    first_pid = list(projects.keys())[0] if projects else None
    for fb in old_fbs:
        prop = fb.get('property', '')
        target_pid = proj_by_prop.get(prop, first_pid)
        if target_pid:
            existing = load_project_feedbacks(target_pid)
            existing.append(fb)
            save_project_feedbacks(target_pid, existing)
    os.rename(fb_file, fb_file + ".bak")


def load_cad_for_project(proj_id: str):
    """保存済みCADファイルを読み込む（クラウド優先・ローカルフォールバック）"""
    if _cloud_ok():
        try:
            val = cloud_store.get(f"proj:{proj_id}:cad")
            if val and val.get("b64"):
                return load_cad_from_bytes(base64.b64decode(val["b64"]))
        except cloud_store.CloudStoreError as exc:
            _warn_cloud_error(exc)
    path = os.path.join(PROJECTS_DIR, f"{proj_id}.txt")
    if os.path.exists(path):
        with open(path, 'rb') as f:
            return load_cad_from_bytes(f.read())
    return None


def delete_project_storage(proj_id: str):
    """案件のローカルファイルとクラウドデータを削除する"""
    for suffix in (".txt", "_data.json", "_feedbacks.json"):
        p = os.path.join(PROJECTS_DIR, f"{proj_id}{suffix}")
        if os.path.exists(p):
            os.remove(p)
    st.session_state.get('_fb_cache', {}).pop(proj_id, None)
    st.session_state.pop(f'_cloud_saved_{proj_id}', None)
    if _cloud_ok():
        try:
            for key_suffix in ("data", "cad", "feedbacks"):
                cloud_store.delete(f"proj:{proj_id}:{key_suffix}")
        except cloud_store.CloudStoreError as exc:
            # 障害検知時点でサーキットブレーカーが働くため以降の削除は中断。
            # クラウド側に残ったデータは再起動後に復活し得る（バナーで通知済み）
            _warn_cloud_error(exc)


def get_project_state(key: str, default=None):
    """現在の案件のsession_state値を取得"""
    proj_id = st.session_state.get('current_project', '')
    proj_key = f"proj_{proj_id}_{key}"
    return st.session_state.get(proj_key, default)


def set_project_state(key: str, value):
    """現在の案件のsession_state値を設定"""
    proj_id = st.session_state.get('current_project', '')
    proj_key = f"proj_{proj_id}_{key}"
    st.session_state[proj_key] = value


def persist_project_mods():
    """現在の案件の修正データをファイルに永続化"""
    proj_id = st.session_state.get('current_project', '')
    if not proj_id:
        return
    data = {}
    mods = get_project_state('mods')
    if mods:
        data['mods'] = {
            'overrides': mods.get('overrides', {}),
            'deleted': mods.get('deleted', set()),
            'new_items': mods.get('new_items', []),
        }
    andpad_mods = get_project_state('andpad_mods')
    if andpad_mods:
        data['andpad_mods'] = andpad_mods
    tp = get_project_state('target_profit_rate')
    if tp is not None:
        data['target_profit_rate'] = tp
    owner = get_project_state('owner_name')
    if owner:
        data['owner_name'] = owner
    loc = get_project_state('location')
    if loc:
        data['location'] = loc
    est_no = get_project_state('estimate_no')
    if est_no:
        data['estimate_no'] = est_no
    inc_adj = get_project_state('include_adjustment')
    if inc_adj is not None:
        data['include_adjustment'] = inc_adj
    save_project_data(proj_id, data)


def restore_project_to_session(proj_id: str):
    """保存済みデータをsession_stateに復元"""
    data = load_project_data(proj_id)
    if 'mods' in data:
        m = data['mods']
        if isinstance(m.get('deleted'), list):
            m['deleted'] = set(m['deleted'])
        st.session_state[f"proj_{proj_id}_mods"] = m
    if 'andpad_mods' in data:
        st.session_state[f"proj_{proj_id}_andpad_mods"] = data['andpad_mods']
    if 'target_profit_rate' in data:
        st.session_state[f"proj_{proj_id}_target_profit_rate"] = data['target_profit_rate']
    if 'owner_name' in data:
        st.session_state[f"proj_{proj_id}_owner_name"] = data['owner_name']
    if 'location' in data:
        st.session_state[f"proj_{proj_id}_location"] = data['location']
    if 'estimate_no' in data:
        st.session_state[f"proj_{proj_id}_estimate_no"] = data['estimate_no']
    if 'include_adjustment' in data:
        st.session_state[f"proj_{proj_id}_include_adjustment"] = data['include_adjustment']
    # CADデータ復元
    cad = load_cad_for_project(proj_id)
    if cad:
        st.session_state[f"proj_{proj_id}_cad_data"] = cad


def _render_project_list_tab(projects: dict, restore_fn):
    """案件一覧タブの中身を描画"""
    st.markdown("#### 案件一覧")

    if not projects:
        st.markdown("")
        st.info("まだ案件がありません。サイドバーからCADデータをアップロードするか「サンプルデータで試す」で始めてください。")
        st.markdown("")
        col1, col2, col3 = st.columns(3, gap="medium")
        with col1:
            st.markdown("""
            <div class="step-card">
                <span class="step-num">STEP 1</span>
                <h3>データ入力</h3>
                <p>ARCHITREND ZEROから出力したCAD数量データ（TXT）をアップロードしてください。サイドバーの「サンプルデータで試す」から動作確認も可能です。</p>
            </div>
            """, unsafe_allow_html=True)
        with col2:
            st.markdown("""
            <div class="step-card">
                <span class="step-num">STEP 2</span>
                <h3>概算見積の自動生成</h3>
                <p>24工事種別の概算見積書を自動で算出します。単価マスタ・計算式マスタに基づき、見積金額・発注金額・粗利を一括計算します。</p>
            </div>
            """, unsafe_allow_html=True)
        with col3:
            st.markdown("""
            <div class="step-card">
                <span class="step-num">STEP 3</span>
                <h3>ANDPAD工務予算出力</h3>
                <p>積算原価を材料費と施工費に分離し、発注先別に集計します。ANDPAD取込用のCSV形式で出力できます。</p>
            </div>
            """, unsafe_allow_html=True)
        return

    current_pid = st.session_state.get('current_project', '')
    st.markdown(f"登録済み **{len(projects)}** 件")
    st.markdown("")

    for pid, pinfo in projects.items():
        is_current = (pid == current_pid)
        border_color = "#2563eb" if is_current else "#e2e8f0"
        bg = "#eff6ff" if is_current else "#ffffff"
        badge = '<span style="background:#2563eb;color:#fff;font-size:0.72rem;font-weight:700;padding:2px 10px;border-radius:4px;margin-left:0.5rem;">表示中</span>' if is_current else ''

        st.markdown(
            f'<div style="background:{bg};border:1.5px solid {border_color};border-radius:10px;'
            f'padding:0.8rem 1.2rem;margin-bottom:0.5rem;box-shadow:0 1px 3px rgba(0,0,0,0.04);">'
            f'<div style="display:flex;align-items:center;justify-content:space-between;">'
            f'<div>'
            f'<span style="font-weight:700;font-size:1rem;color:#1e3a5f;">{pinfo["name"]}</span>{badge}<br>'
            f'<span style="font-size:0.8rem;color:#64748b;">登録: {pinfo.get("created", "-")}　ファイル: {pinfo.get("file_name", "-")}</span>'
            f'</div>'
            f'</div></div>',
            unsafe_allow_html=True,
        )
        col_a, col_b, col_c = st.columns([2, 1, 1])
        with col_b:
            if not is_current:
                if st.button("この案件を開く", key=f"projlist_open_{pid}", use_container_width=True, type="primary"):
                    st.session_state['current_project'] = pid
                    if get_project_state('cad_data') is None:
                        restore_fn(pid)
                    st.rerun()
        with col_c:
            if st.button("削除", key=f"projlist_del_{pid}", use_container_width=True, type="secondary"):
                # CADファイル等（ローカル＋クラウド）も削除
                delete_project_storage(pid)
                for k in list(st.session_state.keys()):
                    if k.startswith(f"proj_{pid}_"):
                        del st.session_state[k]
                del st.session_state['projects'][pid]
                save_project_list(st.session_state['projects'])
                remaining = list(st.session_state['projects'].keys())
                st.session_state['current_project'] = remaining[0] if remaining else ''
                if st.session_state['current_project']:
                    restore_fn(st.session_state['current_project'])
                st.rerun()

    st.markdown("")
    st.markdown(
        '<div class="info-box">新しいCADファイルをアップロードすると新規案件として自動登録されます。'
        'サイドバーのドロップダウンからも案件を切り替えできます。</div>',
        unsafe_allow_html=True,
    )


def main():
    # ヘッダー
    st.markdown("""
    <div class="main-header">
        <h1>日本住建株式会社　見積AIシステム</h1>
        <p>ARCHITREND ZERO 数量データ &rarr; 概算見積書 &rarr; ANDPAD 工務予算（材工分離）</p>
    </div>
    """, unsafe_allow_html=True)

    base_dir = os.path.dirname(os.path.abspath(__file__))

    # クラウド接続障害時はセッション中ずっとバナーを表示する
    if st.session_state.get('_cloud_degraded'):
        st.warning(
            "クラウド保存に接続できないため、一時的にローカル保存で動作しています。"
            "この間の変更はアプリ再起動時に失われる可能性があります。"
            f"（{st.session_state.get('_cloud_error_msg', '')}）"
        )

    # 自動保存：rerunの度に現在の案件データを永続化
    if st.session_state.get('current_project') and st.session_state.get('projects_loaded'):
        persist_project_mods()

    # 案件管理の初期化（ファイルから復元）
    if 'projects_loaded' not in st.session_state:
        st.session_state['projects'] = load_project_list()
        st.session_state['current_project'] = ''
        # 最初の案件を選択して復元（他案件は切替時に遅延復元し、起動時の通信を抑える）
        if st.session_state['projects']:
            first_id = list(st.session_state['projects'].keys())[0]
            st.session_state['current_project'] = first_id
            restore_project_to_session(first_id)
        # 旧グローバルfeedbacks.jsonがあれば案件別に移行
        migrate_global_feedbacks(st.session_state['projects'], base_dir)
        st.session_state['projects_loaded'] = True

    if 'projects' not in st.session_state:
        st.session_state['projects'] = {}
    if 'current_project' not in st.session_state:
        st.session_state['current_project'] = ''

    # サイドバー
    with st.sidebar:
        st.markdown("""
        <div style="text-align:center; padding:0.8rem 0 1rem 0; border-bottom:1px solid #334155; margin-bottom:1rem;">
            <div style="font-size:0.75rem; color:#94a3b8; letter-spacing:0.06em;">日本住建株式会社</div>
            <div style="font-size:1rem; font-weight:700; color:#e2e8f0; margin-top:2px;">見積AIシステム</div>
        </div>
        """, unsafe_allow_html=True)

        # 案件管理セクション
        st.markdown("##### 案件管理")
        projects = st.session_state['projects']

        if projects:
            proj_options = list(projects.keys())
            current_idx = 0
            if st.session_state['current_project'] in proj_options:
                current_idx = proj_options.index(st.session_state['current_project'])
            selected_proj = st.selectbox(
                "案件を選択",
                proj_options,
                index=current_idx,
                format_func=lambda x: projects[x]['name'],
                key="project_selector",
            )
            if selected_proj != st.session_state['current_project']:
                # 現在の案件を保存してから切り替え
                persist_project_mods()
                st.session_state['current_project'] = selected_proj
                # 切り替え先のデータがsession_stateになければ復元
                if get_project_state('cad_data') is None:
                    restore_project_to_session(selected_proj)
                st.rerun()

            # 案件削除
            if st.button("この案件を削除", use_container_width=True, type="secondary"):
                pid = st.session_state['current_project']
                # ファイル削除（ローカル＋クラウド）
                delete_project_storage(pid)
                # session_state削除
                keys_to_del = [k for k in list(st.session_state.keys()) if k.startswith(f"proj_{pid}_")]
                for k in keys_to_del:
                    del st.session_state[k]
                del st.session_state['projects'][pid]
                save_project_list(st.session_state['projects'])
                remaining = list(st.session_state['projects'].keys())
                st.session_state['current_project'] = remaining[0] if remaining else ''
                if st.session_state['current_project']:
                    restore_project_to_session(st.session_state['current_project'])
                st.rerun()

        st.divider()

        st.markdown("##### 物件情報")
        owner_name = st.text_input("施主名", placeholder="例: 山田太郎",
                                    value=get_project_state('owner_name', ''),
                                    key="sidebar_owner")
        location = st.text_input("工事場所", placeholder="例: 愛知県安城市",
                                  value=get_project_state('location', ''),
                                  key="sidebar_location")
        estimate_no = st.text_input("見積番号", placeholder="例: 25-2601-001",
                                     value=get_project_state('estimate_no', ''),
                                     key="sidebar_estimate_no")

        # 入力値を案件stateに保存
        if st.session_state['current_project']:
            set_project_state('owner_name', owner_name)
            set_project_state('location', location)
            set_project_state('estimate_no', estimate_no)

        st.divider()
        st.markdown("##### オプション")
        include_adjustment = st.checkbox(
            "調整分（●33）を含める",
            value=get_project_state('include_adjustment', True),
            help="見積調整項目（建材・下地材/大工手間 各475,000円、発注0円）。実績見積では標準的に計上されます。")
        if st.session_state['current_project']:
            set_project_state('include_adjustment', include_adjustment)

        # 案件保存ボタン
        if st.session_state['current_project']:
            st.divider()
            if st.button("案件データを保存", use_container_width=True, type="primary"):
                persist_project_mods()
                save_project_list(st.session_state['projects'])
                st.toast("案件データを保存しました")

        st.divider()
        use_sample = st.button("サンプルデータで試す",
                                help="古居様邸のデータで動作確認", use_container_width=True)

    # 単価マスタ読み込み
    csv_path = os.path.join(base_dir, "tankamaster_updated.csv")

    if not os.path.exists(csv_path):
        st.error("単価マスタファイル (tankamaster_updated.csv) が見つかりません。")
        return

    prices = load_unit_prices(csv_path)

    # ファイルアップロード
    uploaded_file = st.file_uploader(
        "CAD数量データ（TXTファイル）をアップロード（新規案件として登録）",
        type=['txt', 'TXT'],
        help="ARCHITREND ZEROから出力されたShift-JIS形式の数量データTXTファイル",
    )

    # サンプルデータ処理
    if use_sample:
        sample_path = os.path.join(base_dir, "16-古居徹朗様・古居百恵様邸新築工事【数量データ】.TXT")
        if os.path.exists(sample_path):
            cad_data = load_cad_file(sample_path)
            proj_id = "sample_furui"
            st.session_state['projects'][proj_id] = {
                'name': '古居様邸（サンプル）',
                'file_name': 'サンプルデータ',
                'created': datetime.now().strftime('%Y-%m-%d %H:%M'),
            }
            save_project_list(st.session_state['projects'])
            st.session_state['current_project'] = proj_id
            set_project_state('cad_data', cad_data)
            set_project_state('file_name', "古居様邸サンプルデータ")
            set_project_state('owner_name', '古居徹朗')
            # CADファイルも保存
            with open(sample_path, 'rb') as f:
                save_cad_file(proj_id, f.read())
            persist_project_mods()
            st.rerun()
        else:
            st.warning("サンプルデータファイルが見つかりません。")
            return

    if uploaded_file is not None:
        file_bytes = uploaded_file.read()
        try:
            cad_data = load_cad_from_bytes(file_bytes)
        except Exception:
            st.error("ファイルの読み込みに失敗しました。ARCHITREND ZEROから出力された数量データTXTファイルかご確認ください。")
            st.stop()
        if not cad_data.rooms:
            st.error(
                "部屋データが見つかりませんでした。アップロードされたファイルが【数量データ】のTXTファイルかご確認ください。"
                "（【積算データ】や他の形式のファイルは読み込めません）")
            st.stop()
        # ファイル名ベースのプロジェクトID（安全な文字列に変換）
        safe_name = uploaded_file.name.replace(' ', '_').replace('/', '_').replace('\\', '_')
        proj_id = f"upload_{safe_name}"
        # 同名ファイルが既にあるか確認
        existing = [pid for pid, p in projects.items()
                    if p.get('file_name') == uploaded_file.name]
        if existing:
            proj_id = existing[0]
        st.session_state['projects'][proj_id] = {
            'name': cad_data.property_name or uploaded_file.name,
            'file_name': uploaded_file.name,
            'created': datetime.now().strftime('%Y-%m-%d %H:%M'),
        }
        save_project_list(st.session_state['projects'])
        st.session_state['current_project'] = proj_id
        set_project_state('cad_data', cad_data)
        set_project_state('file_name', uploaded_file.name)
        # CADファイルを永続化
        save_cad_file(proj_id, file_bytes)
        persist_project_mods()
        st.rerun()

    # データがロードされている場合
    cad_data = get_project_state('cad_data')

    # --- 案件一覧タブは常に表示、見積タブはデータがある場合のみ ---
    if cad_data is None:
        # データ未選択: 案件一覧のみ表示
        tab_projects_only, = st.tabs(["案件一覧"])
        with tab_projects_only:
            _render_project_list_tab(projects, restore_project_to_session)
        return

    # 案件バー表示
    if len(st.session_state['projects']) > 1:
        badges = []
        for pid, pinfo in st.session_state['projects'].items():
            cls = "project-badge project-badge-active" if pid == st.session_state['current_project'] else "project-badge"
            badges.append(f'<span class="{cls}">{pinfo["name"]}</span>')
        st.markdown(f'<div class="project-bar">{"".join(badges)}</div>', unsafe_allow_html=True)

    # 見積計算（ベース）
    estimate_base = calculate_estimate(
        cad_data, prices,
        owner_name=owner_name,
        location=location,
        estimate_no=estimate_no,
        include_adjustment=include_adjustment,
    )

    # 修正の初期化と適用
    mods = get_project_state('mods')
    if mods is None:
        mods = {'overrides': {}, 'deleted': set(), 'new_items': []}
        set_project_state('mods', mods)

    estimate = copy.deepcopy(estimate_base)
    if mods['overrides'] or mods['deleted'] or mods['new_items']:
        apply_modifications(estimate, mods)

    # 粗利率調整の適用
    target_profit = get_project_state('target_profit_rate')
    if target_profit is not None:
        target_rate = target_profit / 100.0
        estimate = apply_profit_rate_adjustment(estimate, target_rate)

    # ANDPAD変換
    andpad_budget = convert_to_andpad(estimate)

    # ANDPAD材工分離の修正適用
    andpad_mods = get_project_state('andpad_mods')
    if andpad_mods is None:
        andpad_mods = {}
        set_project_state('andpad_mods', andpad_mods)
    andpad_items_by_key = {it.stable_key: it for it in andpad_budget.items}
    for mod_key, amod in list(andpad_mods.items()):
        item = None
        if mod_key in andpad_items_by_key:
            item = andpad_items_by_key[mod_key]
            # 保存時と明細内容が変わっていたら適用しない（誤適用防止）
            if amod.get('item_name') and amod['item_name'] != item.item_name:
                continue
            if 'base_total' in amod and amod['base_total'] != item.total_cost:
                continue
        elif mod_key.isdigit():
            # 旧・位置インデックスキー（旧バージョンの保存データ）
            idx = int(mod_key)
            if idx < len(andpad_budget.items):
                item = andpad_budget.items[idx]
        if item is None:
            continue

        if 'allocs' in amod:
            item.allocations = [
                ANDPADAllocation(
                    vendor=a['vendor'], kind=a['kind'],
                    amount=int(a['amount']), note=a.get('note', ''),
                )
                for a in amod['allocs']
            ]
        else:
            # 旧形式（material_cost/labor_cost/vendor）からの移行。
            # 旧UIでは明細全体（材・工とも）が指定の発注先に集計されていた
            vendor = amod.get('vendor') or item.vendor
            new_allocs = []
            if amod.get('material_cost', 0) != 0:
                new_allocs.append(ANDPADAllocation(vendor, '材', int(amod['material_cost'])))
            if amod.get('labor_cost', 0) != 0:
                new_allocs.append(ANDPADAllocation(vendor, '工', int(amod['labor_cost'])))
            if not new_allocs:
                new_allocs.append(ANDPADAllocation(vendor, '材', 0))
            item.allocations = new_allocs

    # タブ構成（案件一覧を先頭に追加）
    tab_proj, tab1, tab_mod, tab2, tab3, tab4, tab_fb = st.tabs([
        "案件一覧", "見積概要", "見積修正", "見積明細", "ANDPAD材工分離", "CADデータ確認", "フィードバック"
    ])

    # ===== 案件一覧タブ =====
    with tab_proj:
        _render_project_list_tab(projects, restore_project_to_session)

    # ===== タブ1: 見積概要 =====
    with tab1:
        st.markdown(f"**物件名：** {cad_data.property_name}　／　"
                    f"**構造：** {cad_data.structure}　／　"
                    f"**日付：** {cad_data.date}")

        st.markdown("")

        # メトリクス
        col1, col2, col3, col4 = st.columns(4, gap="medium")
        with col1:
            st.metric("見積金額（税込）", f"¥{estimate.total_estimate_incl_tax:,}")
        with col2:
            st.metric("税抜金額", f"¥{estimate.total_estimate_excl_tax:,}")
        with col3:
            st.metric("延床面積",
                       f"{estimate.total_floor_area_m2:.2f} ㎡",
                       f"{estimate.total_floor_area_tsubo:.2f} 坪")
        with col4:
            rate = estimate.total_profit_rate * 100
            st.metric("粗利率", f"{rate:.1f} %",
                       f"粗利額 ¥{estimate.total_profit:,}")

        st.markdown("")

        # 粗利率調整
        st.markdown("#### 粗利率調整")
        st.markdown(
            '<div class="info-box">目標粗利率を設定すると、元の見積単価の比率を維持しつつ全体の粗利率が目標値になるよう一律調整します。</div>',
            unsafe_allow_html=True,
        )
        adj_col1, adj_col2, adj_col3 = st.columns([2, 1, 1], gap="medium")
        with adj_col1:
            current_rate = get_project_state('target_profit_rate')
            default_val = current_rate if current_rate is not None else round(rate, 1)
            target_rate = st.slider(
                "目標粗利率 (%)",
                min_value=5.0, max_value=50.0,
                value=default_val,
                step=0.5,
                key="profit_rate_slider",
            )
        with adj_col2:
            if st.button("粗利率を適用", type="primary", use_container_width=True):
                set_project_state('target_profit_rate', target_rate)
                st.rerun()
        with adj_col3:
            if st.button("粗利率調整をリセット", use_container_width=True, type="secondary"):
                set_project_state('target_profit_rate', None)
                st.rerun()

        if get_project_state('target_profit_rate') is not None:
            st.info(f"目標粗利率 {get_project_state('target_profit_rate'):.1f}% で調整中です。リセットボタンで元の見積に戻ります。")

        st.markdown("")

        # 警告メッセージ
        if estimate.warnings:
            st.markdown("#### 確認事項")
            for warning in estimate.warnings:
                st.markdown(f'<div class="warning-box">{warning}</div>',
                           unsafe_allow_html=True)
            st.markdown("")

        # 工事別内訳表
        st.markdown("#### 工事別内訳")

        summary_data = []
        for cat in estimate.categories:
            profit_rate = f"{cat.profit_rate * 100:.1f}%" if cat.estimate_total > 0 else "-"
            summary_data.append({
                "No": cat.no,
                "工事種別": cat.name,
                "見積金額": cat.estimate_total,
                "発注金額": cat.order_total,
                "粗利額": cat.profit,
                "粗利率": profit_rate,
            })

        df_summary = pd.DataFrame(summary_data)
        st.dataframe(
            df_summary.style.format({
                "見積金額": "¥{:,.0f}",
                "発注金額": "¥{:,.0f}",
                "粗利額": "¥{:,.0f}",
            }),
            use_container_width=True,
            hide_index=True,
            height=700,
        )

        # 工事別グラフ
        st.markdown("#### 工事別構成比")
        chart_data = pd.DataFrame({
            '工事種別': [c.name for c in estimate.categories if c.estimate_total > 0],
            '見積金額': [c.estimate_total for c in estimate.categories if c.estimate_total > 0],
        })
        st.bar_chart(chart_data.set_index('工事種別'), height=400)

        # ダウンロード
        st.markdown("")
        st.markdown("#### ファイル出力")

        file_label = f"{owner_name}様邸" if owner_name else "概算"

        st.markdown("##### Excel出力")
        ex_col1, ex_col2 = st.columns(2, gap="medium")
        with ex_col1:
            excel_data = create_excel_estimate(estimate)
            st.download_button(
                label="社内用見積書（Excel）",
                data=excel_data,
                file_name=f"{file_label}_社内見積.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="secondary",
                use_container_width=True,
            )
        with ex_col2:
            client_excel = create_client_excel(estimate)
            st.download_button(
                label="お客様提出用見積書（Excel）",
                data=client_excel,
                file_name=f"{file_label}_お客様提出用見積書.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )

        st.markdown("##### CSV出力")
        csv_col1, csv_col2, csv_col3, csv_col4 = st.columns(4, gap="small")
        with csv_col1:
            st.download_button(
                label="工事別内訳（CSV）",
                data=create_summary_csv(estimate),
                file_name=f"{file_label}_工事別内訳.csv",
                mime="text/csv",
                use_container_width=True,
            )
        with csv_col2:
            st.download_button(
                label="見積明細（CSV）",
                data=create_detail_csv(estimate),
                file_name=f"{file_label}_見積明細.csv",
                mime="text/csv",
                use_container_width=True,
            )
        with csv_col3:
            st.download_button(
                label="発注先別集計（CSV）",
                data=create_vendor_csv(andpad_budget),
                file_name=f"{file_label}_発注先別集計.csv",
                mime="text/csv",
                use_container_width=True,
            )
        with csv_col4:
            st.download_button(
                label="ANDPAD工務予算（CSV）",
                data=create_andpad_csv(andpad_budget),
                file_name=f"{file_label}_ANDPAD工務予算.csv",
                mime="text/csv",
                use_container_width=True,
            )

    # ===== 見積修正タブ =====
    with tab_mod:
        st.markdown("#### 見積修正入力")
        st.markdown(
            '<div class="warning-box">お客様からの修正依頼をここで入力してください。'
            '工事種別を開いて数量・単価を変更し「変更を適用」を押すと、見積全体が自動再計算されます。</div>',
            unsafe_allow_html=True,
        )
        st.markdown("")

        mod_cat_options = [f"{c.no}. {c.name}" for c in estimate_base.categories if c.no != 22]

        # --- 全工事種別を一覧表示 ---
        for base_cat in estimate_base.categories:
            if base_cat.no == 22:
                continue
            if not base_cat.items:
                continue

            # この工事に修正があるか判定
            cat_mod_count = sum(
                1 for k in mods['overrides'] if k.startswith(f"{base_cat.no}_")
            ) + sum(
                1 for k in mods['deleted'] if k.startswith(f"{base_cat.no}_")
            ) + sum(
                1 for ni in mods['new_items'] if ni['cat_no'] == base_cat.no
            )

            # 現在の小計（修正適用後）
            mod_cat = None
            for c in estimate.categories:
                if c.no == base_cat.no:
                    mod_cat = c
                    break
            cat_est_total = mod_cat.estimate_total if mod_cat else 0
            cat_ord_total = mod_cat.order_total if mod_cat else 0

            mod_badge = f"  [{cat_mod_count}件修正中]" if cat_mod_count > 0 else ""
            exp_label = (
                f"{base_cat.no}. {base_cat.name}　—　"
                f"見積: ¥{cat_est_total:,} / 発注: ¥{cat_ord_total:,}"
                f"{mod_badge}"
            )

            with st.expander(exp_label, expanded=False):
                edit_rows = []
                for i, item in enumerate(base_cat.items):
                    key = f"{base_cat.no}_{i}"
                    is_del = key in mods['deleted']
                    ov = mods['overrides'].get(key, {})
                    edit_rows.append({
                        '削除': is_del,
                        '名称': item.name,
                        '摘要': item.summary,
                        '数量': ov.get('quantity', item.quantity),
                        '単位': item.unit,
                        '見積単価': ov.get('estimate_price', item.estimate_price),
                        '発注単価': ov.get('order_price', item.order_price),
                    })

                df_edit = pd.DataFrame(edit_rows)

                edited = st.data_editor(
                    df_edit,
                    disabled=['単位'],
                    use_container_width=True,
                    hide_index=True,
                    key=f"editor_{base_cat.no}",
                    column_config={
                        '削除': st.column_config.CheckboxColumn("削除", default=False),
                        '名称': st.column_config.TextColumn("名称", help="項目名を直接編集できます"),
                        '摘要': st.column_config.TextColumn("摘要", help="摘要を直接編集できます"),
                        '数量': st.column_config.NumberColumn("数量", format="%.2f", min_value=0.0),
                        '見積単価': st.column_config.NumberColumn("見積単価", min_value=0),
                        '発注単価': st.column_config.NumberColumn("発注単価", min_value=0),
                    },
                )

                col_a, col_b = st.columns(2)
                with col_a:
                    if st.button("変更を適用", type="primary",
                                 use_container_width=True, key=f"apply_{base_cat.no}"):
                        for i, item in enumerate(base_cat.items):
                            key = f"{base_cat.no}_{i}"
                            row = edited.iloc[i]
                            if row['削除']:
                                mods['deleted'].add(key)
                                mods['overrides'].pop(key, None)
                            else:
                                mods['deleted'].discard(key)
                                new_name = str(row['名称'])
                                new_summary = str(row['摘要']) if row['摘要'] is not None else ''
                                changed = (
                                    abs(float(row['数量']) - item.quantity) > 0.001
                                    or int(row['見積単価']) != item.estimate_price
                                    or int(row['発注単価']) != item.order_price
                                    or new_name != item.name
                                    or new_summary != (item.summary or '')
                                )
                                if changed:
                                    mods['overrides'][key] = {
                                        'quantity': float(row['数量']),
                                        'estimate_price': int(row['見積単価']),
                                        'order_price': int(row['発注単価']),
                                        'name': new_name,
                                        'summary': new_summary,
                                    }
                                else:
                                    mods['overrides'].pop(key, None)
                        st.rerun()

                with col_b:
                    if st.button("リセット", use_container_width=True,
                                 key=f"reset_{base_cat.no}", type="secondary"):
                        to_del = [k for k in mods['overrides']
                                  if k.startswith(f"{base_cat.no}_")]
                        for k in to_del:
                            del mods['overrides'][k]
                        mods['deleted'] = {
                            k for k in mods['deleted']
                            if not k.startswith(f"{base_cat.no}_")
                        }
                        editor_key = f"editor_{base_cat.no}"
                        if editor_key in st.session_state:
                            del st.session_state[editor_key]
                        st.rerun()

        st.divider()

        # 新規項目追加
        st.markdown("#### 新規項目の追加")
        with st.form("new_item_form"):
            ni_cat = st.selectbox("追加先の工事種別", mod_cat_options, key="ni_cat")
            col1, col2 = st.columns(2)
            with col1:
                ni_name = st.text_input("名称", placeholder="例: 追加オプション工事")
            with col2:
                ni_summary = st.text_input("摘要", placeholder="例: お客様要望による追加")
            col1, col2, col3 = st.columns(3)
            with col1:
                ni_qty = st.number_input("数量", value=1.0, min_value=0.01, format="%.2f")
            with col2:
                ni_unit = st.text_input("単位", value="式")
            with col3:
                pass
            col1, col2 = st.columns(2)
            with col1:
                ni_ep = st.number_input("見積単価（円）", value=0, min_value=0)
            with col2:
                ni_op = st.number_input("発注単価（円）", value=0, min_value=0)
            if st.form_submit_button("項目を追加"):
                if ni_name:
                    mods['new_items'].append({
                        'cat_no': int(ni_cat.split(".")[0]),
                        'name': ni_name,
                        'summary': ni_summary,
                        'quantity': ni_qty,
                        'unit': ni_unit,
                        'estimate_price': ni_ep,
                        'order_price': ni_op,
                    })
                    st.rerun()

        # 追加済み項目一覧
        if mods['new_items']:
            st.markdown("##### 追加済み項目")
            for idx, ni in enumerate(mods['new_items']):
                cat_name = ""
                for c in estimate_base.categories:
                    if c.no == ni['cat_no']:
                        cat_name = c.name
                        break
                col1, col2 = st.columns([5, 1])
                with col1:
                    est_amt = int(ni['quantity'] * ni['estimate_price'])
                    st.markdown(
                        f"**{cat_name}** > {ni['name']}"
                        f"（{ni['quantity']} {ni['unit']}）— 見積: ¥{est_amt:,}"
                    )
                with col2:
                    if st.button("削除", key=f"del_new_{idx}", type="secondary"):
                        mods['new_items'].pop(idx)
                        st.rerun()

        # 修正サマリ
        mods_summary = mods
        has_any_mods = mods_summary['overrides'] or mods_summary['deleted'] or mods_summary['new_items']
        if has_any_mods:
            st.divider()
            st.markdown("#### 修正サマリ")
            mod_count = (len(mods_summary['overrides'])
                         + len(mods_summary['deleted'])
                         + len(mods_summary['new_items']))
            st.info(f"合計 {mod_count} 件の修正が適用されています。")

            if mods_summary['overrides']:
                st.markdown("**単価・数量変更:**")
                for key, ov in mods_summary['overrides'].items():
                    parts = key.split('_')
                    c_no, c_idx = int(parts[0]), int(parts[1])
                    for c in estimate_base.categories:
                        if c.no == c_no and c_idx < len(c.items):
                            orig = c.items[c_idx]
                            changes = []
                            if abs(ov['quantity'] - orig.quantity) > 0.001:
                                changes.append(f"数量: {orig.quantity} → {ov['quantity']}")
                            if ov['estimate_price'] != orig.estimate_price:
                                changes.append(
                                    f"見積単価: ¥{orig.estimate_price:,} → ¥{ov['estimate_price']:,}"
                                )
                            if ov['order_price'] != orig.order_price:
                                changes.append(
                                    f"発注単価: ¥{orig.order_price:,} → ¥{ov['order_price']:,}"
                                )
                            if changes:
                                st.markdown(f"- **{c.name}** > {orig.name}: {', '.join(changes)}")

            if mods_summary['deleted']:
                st.markdown("**削除項目:**")
                for key in mods_summary['deleted']:
                    parts = key.split('_')
                    c_no, c_idx = int(parts[0]), int(parts[1])
                    for c in estimate_base.categories:
                        if c.no == c_no and c_idx < len(c.items):
                            st.markdown(f"- ~~{c.name} > {c.items[c_idx].name}~~")

            st.markdown("")
            if st.button("全修正をリセット", type="secondary"):
                set_project_state('mods', {'overrides': {}, 'deleted': set(), 'new_items': []})
                keys_to_clear = [k for k in list(st.session_state.keys()) if k.startswith("editor_")]
                for k in keys_to_clear:
                    del st.session_state[k]
                st.rerun()

    # ===== タブ2: 見積明細 =====
    with tab2:
        st.markdown("#### 見積明細書")

        # 工事種別フィルター
        cat_names = ["全て"] + [f"{c.no}. {c.name}" for c in estimate.categories]
        selected_cat = st.selectbox("工事種別フィルター", cat_names)

        detail_data = []
        for cat in estimate.categories:
            if selected_cat != "全て" and f"{cat.no}. {cat.name}" != selected_cat:
                continue
            for item in cat.items:
                detail_data.append({
                    "工事種別": f"{cat.no}. {cat.name}",
                    "カテゴリ": item.category or "-",
                    "名称": item.name,
                    "摘要": item.summary,
                    "数量": item.quantity,
                    "単位": item.unit,
                    "見積単価": item.estimate_price,
                    "見積金額": item.estimate_amount,
                    "発注単価": item.order_price,
                    "発注金額": item.order_amount,
                })

        if detail_data:
            df_detail = pd.DataFrame(detail_data)
            st.dataframe(
                df_detail.style.format({
                    "数量": "{:.2f}",
                    "見積単価": "¥{:,.0f}",
                    "見積金額": "¥{:,.0f}",
                    "発注単価": "¥{:,.0f}",
                    "発注金額": "¥{:,.0f}",
                }),
                use_container_width=True,
                hide_index=True,
                height=600,
            )

            # 選択中カテゴリの小計
            if selected_cat != "全て":
                total_est = sum(d["見積金額"] for d in detail_data)
                total_ord = sum(d["発注金額"] for d in detail_data)
                st.markdown(f"**小計 - 見積: ¥{total_est:,} / 発注: ¥{total_ord:,}**")

            st.markdown("")
            file_label_d = f"{owner_name}様邸" if owner_name else "概算"
            st.download_button(
                label="見積明細をCSVダウンロード",
                data=create_detail_csv(estimate),
                file_name=f"{file_label_d}_見積明細.csv",
                mime="text/csv",
            )

    # ===== タブ3: ANDPAD材工分離 =====
    with tab3:
        st.markdown("#### ANDPAD工務予算 ― 材工分離")
        st.markdown("積算原価を材料費と施工費に分離し、発注先（細目工種）別に集計します。")

        # 集計方法の選択
        view_mode = st.radio("表示モード",
                              ["発注先別集計", "工事種別別集計", "明細一覧（編集可）"],
                              horizontal=True)

        if view_mode == "発注先別集計":
            by_vendor = andpad_budget.by_vendor()

            st.markdown(
                '<div class="warning-box">発注先をクリックすると明細が開閉します。'
                'ANDPADでの発注時は、明細を開いてコピー＆ペーストしてください。'
                '並び順は予実管理表（積算原価⇒工務予算振り分け方法）に合わせています。</div>',
                unsafe_allow_html=True,
            )
            st.markdown("")

            for vendor, data in by_vendor.items():
                exp_label = (
                    f"{vendor}　—　材料: ¥{data['material']:,} / "
                    f"施工: ¥{data['labor']:,} / 合計: ¥{data['total']:,}"
                    f"（{len(data['details'])}件）"
                )
                with st.expander(exp_label, expanded=False):
                    detail_rows = []
                    for item, alloc in data['details']:
                        summary = item.summary
                        if len(item.allocations) > 1 and alloc.note:
                            summary = (f"{summary}　※{alloc.note}"
                                       if summary else f"※{alloc.note}")
                        detail_rows.append({
                            "工事種別": item.work_category,
                            "名称": item.item_name,
                            "摘要": summary,
                            "数量": item.quantity,
                            "単位": item.unit,
                            "材料費": alloc.amount if alloc.kind == '材' else 0,
                            "施工費": alloc.amount if alloc.kind == '工' else 0,
                            "金額": alloc.amount,
                        })
                    df_detail = pd.DataFrame(detail_rows)
                    st.dataframe(
                        df_detail.style.format({
                            "数量": "{:,.2f}",
                            "材料費": "¥{:,.0f}",
                            "施工費": "¥{:,.0f}",
                            "金額": "¥{:,.0f}",
                        }),
                        use_container_width=True,
                        hide_index=True,
                    )

        elif view_mode == "工事種別別集計":
            by_cat = andpad_budget.by_work_category()
            cat_summary = []
            for cat_name, data in by_cat.items():
                material_ratio = data['material'] / data['total'] * 100 if data['total'] > 0 else 0
                cat_summary.append({
                    "工事種別": cat_name,
                    "材料費": data['material'],
                    "施工費": data['labor'],
                    "合計": data['total'],
                    "材料比率": f"{material_ratio:.1f}%",
                })

            df_cat = pd.DataFrame(cat_summary)
            st.dataframe(
                df_cat.style.format({
                    "材料費": "¥{:,.0f}",
                    "施工費": "¥{:,.0f}",
                    "合計": "¥{:,.0f}",
                }),
                use_container_width=True,
                hide_index=True,
                height=600,
            )

        else:  # 明細一覧（編集可）
            st.markdown(
                '<div class="warning-box">材工分離した明細は発注先ごとに1行ずつ表示されます'
                '（例: 材料費→建材、施工費→大工手間）。発注先・材料費・施工費を編集し、'
                '「変更を適用」ボタンを押してください。'
                '施工費の行の発注先を変えると、その発注先の集計に加算されます。</div>',
                unsafe_allow_html=True,
            )
            st.markdown("")

            # 工事種別でグルーピングして表示
            cat_items = {}  # cat_name -> [(global_idx, item)]
            for g_idx, item in enumerate(andpad_budget.items):
                cat_name = item.work_category
                if cat_name not in cat_items:
                    cat_items[cat_name] = []
                cat_items[cat_name].append((g_idx, item))

            for cat_name, items_list in cat_items.items():
                cat_material = sum(it.material_cost for _, it in items_list)
                cat_labor = sum(it.labor_cost for _, it in items_list)
                cat_total = sum(it.total_cost for _, it in items_list)
                mod_count = sum(1 for g_idx, it in items_list
                                if it.stable_key in andpad_mods or str(g_idx) in andpad_mods)
                mod_badge = f"  [{mod_count}件修正中]" if mod_count > 0 else ""

                exp_label = (
                    f"{cat_name}　—　"
                    f"材料: ¥{cat_material:,} / 施工: ¥{cat_labor:,} / 計: ¥{cat_total:,}"
                    f"{mod_badge}"
                )

                with st.expander(exp_label, expanded=False):
                    # 配分（材工分離後）単位で1行ずつ
                    edit_rows = []
                    row_keys = []  # (g_idx, alloc_idx, kind, note)
                    for g_idx, item in items_list:
                        multi = len(item.allocations) > 1
                        for a_idx, alloc in enumerate(item.allocations):
                            summary = item.summary
                            if multi and alloc.note:
                                summary = (f"{summary}　※{alloc.note}"
                                           if summary else f"※{alloc.note}")
                            edit_rows.append({
                                '発注先': alloc.vendor,
                                '名称': item.item_name,
                                '摘要': summary,
                                '数量': item.quantity,
                                '単位': item.unit,
                                '材料費': alloc.amount if alloc.kind == '材' else 0,
                                '施工費': alloc.amount if alloc.kind == '工' else 0,
                            })
                            row_keys.append((g_idx, a_idx, alloc.kind, alloc.note))

                    df_edit = pd.DataFrame(edit_rows)
                    edited = st.data_editor(
                        df_edit,
                        key=f"andpad_editor_{cat_name}",
                        use_container_width=True,
                        disabled=['名称', '摘要', '数量', '単位'],
                        column_config={
                            '発注先': st.column_config.TextColumn('発注先', width='medium'),
                            '名称': st.column_config.TextColumn('名称', width='medium'),
                            '摘要': st.column_config.TextColumn('摘要', width='medium'),
                            '数量': st.column_config.NumberColumn('数量', format="%.2f"),
                            '単位': st.column_config.TextColumn('単位', width='small'),
                            '材料費': st.column_config.NumberColumn('材料費', format="¥%d"),
                            '施工費': st.column_config.NumberColumn('施工費', format="¥%d"),
                        },
                        hide_index=True,
                    )

                    a_col1, a_col2 = st.columns([1, 1])
                    with a_col1:
                        if st.button("変更を適用", key=f"andpad_apply_{cat_name}", type="primary"):
                            # 編集結果から明細ごとの配分リストを再構築
                            new_allocs_by_item = {}  # g_idx -> [alloc dict]
                            for row_i, (g_idx, a_idx, kind, note) in enumerate(row_keys):
                                raw_vendor = edited.iloc[row_i]['発注先']
                                raw_mat = edited.iloc[row_i]['材料費']
                                raw_lab = edited.iloc[row_i]['施工費']
                                vendor = ('' if pd.isna(raw_vendor)
                                          else str(raw_vendor).strip())
                                if not vendor:
                                    vendor = andpad_budget.items[g_idx].vendor
                                mat = 0 if pd.isna(raw_mat) else int(raw_mat)
                                lab = 0 if pd.isna(raw_lab) else int(raw_lab)
                                allocs = new_allocs_by_item.setdefault(g_idx, [])
                                if mat != 0:
                                    allocs.append({'vendor': vendor, 'kind': '材',
                                                   'amount': mat,
                                                   'note': note if kind == '材' else ''})
                                if lab != 0:
                                    allocs.append({'vendor': vendor, 'kind': '工',
                                                   'amount': lab,
                                                   'note': note if kind == '工' else ''})

                            # 元のANDPADデータ（修正前）と比較して差分のみ保存
                            orig_budget = convert_to_andpad(estimate)
                            orig_by_key = {it.stable_key: it
                                           for it in orig_budget.items}
                            for g_idx, cur_item in items_list:
                                mod_key = cur_item.stable_key
                                orig_item = orig_by_key.get(mod_key)
                                if orig_item is None:
                                    continue
                                new_allocs = new_allocs_by_item.get(g_idx, [])
                                if not new_allocs:
                                    # 全て0にされた場合も明細は残す（金額0・元の区分を維持）
                                    base = (orig_item.allocations[0]
                                            if orig_item.allocations else None)
                                    new_allocs = [{
                                        'vendor': base.vendor if base else orig_item.vendor,
                                        'kind': base.kind if base else '材',
                                        'amount': 0, 'note': '',
                                    }]
                                # 金額0の配分は比較から除外（幽霊修正の防止）
                                orig_sig = sorted((a.vendor, a.kind, a.amount)
                                                  for a in orig_item.allocations
                                                  if a.amount != 0)
                                new_sig = sorted((a['vendor'], a['kind'], a['amount'])
                                                 for a in new_allocs
                                                 if a['amount'] != 0)
                                if new_sig != orig_sig:
                                    andpad_mods[mod_key] = {
                                        'allocs': new_allocs,
                                        'item_name': orig_item.item_name,
                                        'base_total': orig_item.total_cost,
                                    }
                                else:
                                    andpad_mods.pop(mod_key, None)
                                # 旧・位置インデックスキーの残骸は掃除
                                andpad_mods.pop(str(g_idx), None)
                            st.rerun()
                    with a_col2:
                        if st.button("この工事のリセット", key=f"andpad_reset_{cat_name}", type="secondary"):
                            for g_idx, cur_item in items_list:
                                andpad_mods.pop(cur_item.stable_key, None)
                                andpad_mods.pop(str(g_idx), None)
                            key = f"andpad_editor_{cat_name}"
                            if key in st.session_state:
                                del st.session_state[key]
                            st.rerun()

        # 材工比率サマリ
        st.markdown("")
        st.markdown("#### 材工比率")
        col1, col2, col3 = st.columns(3, gap="medium")
        with col1:
            st.metric("材料費合計", f"¥{andpad_budget.total_material:,}")
        with col2:
            st.metric("施工費合計", f"¥{andpad_budget.total_labor:,}")
        with col3:
            st.metric("発注総額", f"¥{andpad_budget.grand_total:,}")

        ratio_data = pd.DataFrame({
            '区分': ['材料費', '施工費'],
            '金額': [andpad_budget.total_material, andpad_budget.total_labor],
        })
        st.bar_chart(ratio_data.set_index('区分'), height=300)

        # ANDPAD CSV ダウンロード
        st.markdown("")
        st.markdown("#### ANDPAD用CSVダウンロード")
        andpad_dl_col1, andpad_dl_col2 = st.columns([2, 2])
        with andpad_dl_col1:
            andpad_csv_dl = create_andpad_csv(andpad_budget)
            file_label_andpad = f"{owner_name}様邸" if owner_name else "概算"
            st.download_button(
                label="ANDPAD工務予算（CSV）をダウンロード",
                data=andpad_csv_dl,
                file_name=f"{file_label_andpad}_ANDPAD工務予算.csv",
                mime="text/csv",
                type="primary",
                use_container_width=True,
            )
        with andpad_dl_col2:
            if andpad_mods:
                st.info(f"{len(andpad_mods)} 件の材工修正が適用されています。")
                if st.button("全材工修正をリセット", type="secondary"):
                    set_project_state('andpad_mods', {})
                    keys_to_clear = [k for k in list(st.session_state.keys()) if k.startswith("andpad_editor_")]
                    for k in keys_to_clear:
                        del st.session_state[k]
                    st.rerun()

    # ===== タブ4: CADデータ確認 =====
    with tab4:
        st.markdown("#### CADデータ解析結果")

        col1, col2 = st.columns(2)

        with col1:
            st.markdown("##### 物件情報")
            info_data = {
                "物件番号": cad_data.property_number,
                "物件名": cad_data.property_name,
                "図面名": cad_data.drawing_name,
                "担当者": cad_data.designer,
                "日付": cad_data.date,
                "構造": cad_data.structure,
            }
            st.json(info_data)

            st.markdown("##### 面積情報")
            area_data = {
                "延床面積（㎡）": f"{cad_data.total_floor_area():.2f}",
                "延床面積（坪）": f"{cad_data.total_floor_area_tsubo():.2f}",
            }
            for floor in cad_data.get_floor_names():
                area = cad_data.floor_area(floor)
                area_data[f"{floor}面積（㎡）"] = f"{area:.2f}"
            st.json(area_data)

        with col2:
            st.markdown("##### 特長条件")
            cond_data = []
            cond_meanings = {
                "R000051": "鋼板屋根",
                "R000052": "準防火地域",
                "R000053": "IH採用",
                "R000054": "太陽光発電",
            }
            for cond in cad_data.conditions:
                if cond.code in cond_meanings:
                    cond_data.append({
                        "コード": cond.code,
                        "条件名": cond_meanings.get(cond.code, cond.name),
                        "値": "ON" if cond.value == 1 else "OFF",
                    })
            if cond_data:
                st.dataframe(pd.DataFrame(cond_data), hide_index=True)

            st.markdown("##### 部屋一覧")
            room_data = []
            for room in cad_data.rooms:
                room_data.append({
                    "階": room.floor,
                    "部屋名": room.name,
                    "床面積(㎡)": f"{room.quantities.get('N000001', 0):.2f}",
                    "天井面積(㎡)": f"{room.quantities.get('N000003', 0):.2f}",
                    "天井高(m)": f"{room.quantities.get('N000050', 0):.2f}",
                })
            if room_data:
                st.dataframe(pd.DataFrame(room_data),
                            use_container_width=True, hide_index=True, height=400)

        st.markdown("##### 建具一覧")
        fitting_data = []
        for f in cad_data.fittings:
            fitting_data.append({
                "種別": f.category,
                "形式": f.type_name,
                "材質": f.material,
                "コード": f.code,
                "数量": int(f.quantities.get("T100001", 0)),
                "幅(m)": f"{f.quantities.get('T100010', 0):.3f}",
                "高(m)": f"{f.quantities.get('T100012', 0):.3f}",
            })
        if fitting_data:
            st.dataframe(pd.DataFrame(fitting_data),
                        use_container_width=True, hide_index=True)

    # ===== フィードバックタブ =====
    with tab_fb:
        current_pid = st.session_state.get('current_project', '')
        current_pname = st.session_state['projects'].get(current_pid, {}).get('name', '') if current_pid else ''

        st.markdown("#### フィードバック・ご要望")
        st.markdown(
            '<div class="warning-box">'
            f'現在の案件: <b>{current_pname or "未選択"}</b><br>'
            'お客様や担当者からのご意見・ご要望をここから送信してください。<br>'
            '送信された内容は<b>案件ごとに</b>ファイルに保存され、開発チームが確認のうえシステムに反映します。<br>'
            '「単価が実態と違う」「出力フォーマットを変えてほしい」など、何でもお気軽にどうぞ。'
            '</div>',
            unsafe_allow_html=True,
        )
        st.markdown("")

        # 案件ごとのフィードバック読み込み
        if current_pid:
            feedbacks = load_project_feedbacks(current_pid)
        else:
            feedbacks = []

        # 送信フォーム
        with st.form("feedback_form", clear_on_submit=True):
            fb_category = st.selectbox(
                "カテゴリ",
                [
                    "単価について（実態と合わない等）",
                    "数量・計算ロジックについて",
                    "出力フォーマットについて（Excel等）",
                    "項目の追加・削除要望",
                    "ANDPAD連携について",
                    "操作性・画面について",
                    "その他",
                ],
                key="fb_category",
            )

            fb_target = st.text_input(
                "対象（任意）",
                placeholder="例: 03.木工事 > プレカット / キッチンの単価 / Excel出力",
            )

            fb_body = st.text_area(
                "内容",
                placeholder=(
                    "例:\n"
                    "・プレカットの見積単価が実際の仕入れ値より高い気がします。現在の相場は坪○○円くらいです。\n"
                    "・外壁サイディングの種類別（窯業系/金属系）に単価を分けてほしい。\n"
                    "・見積書のExcelに備考欄を追加してほしい。"
                ),
                height=150,
            )

            fb_priority = st.radio(
                "優先度",
                ["通常", "急ぎ"],
                horizontal=True,
            )

            fb_name = st.text_input(
                "送信者名（任意）",
                placeholder="例: 山田",
            )

            fb_email = st.text_input(
                "メールアドレス（任意）",
                placeholder="返信が必要な場合のみご入力ください",
            )

            submitted = st.form_submit_button("フィードバックを送信", type="primary")
            if submitted and fb_body:
                if not current_pid:
                    st.error("案件が選択されていません。先にCADデータをアップロードしてください。")
                else:
                    property_name_for_fb = cad_data.property_name if cad_data else ''
                    feedbacks.append({
                        'category': fb_category,
                        'target': fb_target,
                        'body': fb_body,
                        'priority': fb_priority,
                        'name': fb_name,
                        'email': fb_email,
                        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M'),
                        'property': property_name_for_fb,
                        'status': '未対応',
                    })
                    save_project_feedbacks(current_pid, feedbacks)

                    # 弊社のSlackチャンネルへ即時通知（メール併設定時はメールも）
                    with st.spinner("フィードバックを弊社に送信しています..."):
                        send_result = send_feedback_to_company(
                            name=fb_name or "(匿名)",
                            email=fb_email or "",
                            category=fb_category,
                            message=fb_body,
                            property_name=property_name_for_fb,
                            priority=fb_priority,
                            target=fb_target,
                        )

                    if send_result.get("success"):
                        channels = "・".join(
                            {"slack": "Slack", "email": "メール", "local_json": "ローカル"}.get(c, c)
                            for c in send_result.get("channels_sent", [])
                        )
                        st.success(
                            f"フィードバックを送信しました（通知先: {channels}）。"
                            f" 受付ID: `{send_result['feedback_id'][:8]}` 担当者が確認後、必要に応じてご連絡いたします。"
                        )
                        st.balloons()
                    else:
                        errs = "; ".join(send_result.get("errors", [])) or "不明なエラー"
                        st.warning(
                            f"案件フォルダへの保存は完了しましたが、弊社への通知に一部失敗しました: {errs}"
                        )
                    st.rerun()

        # 仕組みの説明
        st.markdown("")
        st.markdown("##### このフィードバック機能について")
        st.markdown(
            "- 送信されたフィードバックは **案件ごとにサーバー上のファイルに保存** されます（ブラウザを閉じても消えません）\n"
            "- 開発チームがフィードバックを確認し、**単価マスタの修正・計算ロジックの改善・画面変更** などをシステムに反映します\n"
            "- 対応状況はこの画面で確認できます（未対応 → 対応済み）\n"
            "- 「全案件横断表示」で全案件のフィードバックをまとめて確認できます\n"
            "- フィードバック一覧はCSVでダウンロードして共有することもできます"
        )

        # 表示モード切替
        st.divider()
        fb_view_mode = st.radio(
            "表示範囲",
            ["この案件のみ", "全案件横断表示"],
            horizontal=True,
            key="fb_view_mode",
        )

        if fb_view_mode == "全案件横断表示":
            display_fbs = load_all_feedbacks(st.session_state['projects'])
            show_project_name = True
            st.markdown(f"##### 全案件のフィードバック一覧")
        else:
            display_fbs = feedbacks
            show_project_name = False
            if current_pname:
                st.markdown(f"##### 「{current_pname}」のフィードバック")

        # 送信済みフィードバック一覧
        if display_fbs:
            # ステータスフィルター
            status_options = ["全て", "未対応", "対応済み"]
            fb_filter = st.radio("表示フィルター", status_options, horizontal=True, key="fb_filter")

            filtered = display_fbs if fb_filter == "全て" else [
                fb for fb in display_fbs if fb.get('status', '未対応') == fb_filter
            ]

            undone_count = sum(1 for fb in display_fbs if fb.get('status', '未対応') == '未対応')
            done_count = sum(1 for fb in display_fbs if fb.get('status') == '対応済み')
            st.markdown(f"全 {len(display_fbs)} 件（未対応: {undone_count} / 対応済み: {done_count}）")

            for idx, fb in enumerate(filtered if fb_view_mode == "全案件横断表示" else list(reversed(filtered))):
                if fb_view_mode == "全案件横断表示":
                    fb_proj_id = fb.get('_proj_id', '')
                    fb_proj_name = fb.get('_proj_name', '')
                else:
                    fb_proj_id = current_pid
                    fb_proj_name = current_pname
                    # 元のインデックスを特定
                actual_idx_in_list = None
                if fb_view_mode != "全案件横断表示":
                    try:
                        actual_idx_in_list = len(feedbacks) - 1 - list(reversed(feedbacks)).index(fb)
                    except ValueError:
                        actual_idx_in_list = idx

                priority_mark = '<span style="color:#dc2626;font-weight:700;">【急ぎ】</span>' if fb.get('priority') == '急ぎ' else ""
                status = fb.get('status', '未対応')
                status_color = "#dc2626" if status == "未対応" else "#16a34a"
                status_html = f'<span style="color:{status_color};font-weight:600;">[{status}]</span>'
                proj_label = f'<span style="background:#e0f2fe;color:#0c4a6e;font-size:0.75rem;font-weight:600;padding:2px 8px;border-radius:4px;margin-right:0.4rem;">{fb_proj_name}</span>' if show_project_name and fb_proj_name else ""
                st.markdown(
                    f'<div class="feedback-card">'
                    f'<div class="fb-meta">'
                    f'{proj_label}'
                    f'<span class="fb-category">{fb["category"]}</span>'
                    f'{status_html} {priority_mark} '
                    f'{fb["timestamp"]}'
                    f'{" / " + fb["name"] if fb.get("name") else ""}'
                    f'{" / " + fb["property"] if fb.get("property") else ""}'
                    f'</div>'
                    f'<div class="fb-body">'
                    f'{"<b>対象:</b> " + fb["target"] + "<br>" if fb.get("target") else ""}'
                    f'{fb["body"]}'
                    f'</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

                # ステータス変更・削除ボタン（案件横断表示でも操作可能）
                target_pid = fb.get('_proj_id', current_pid) if fb_view_mode == "全案件横断表示" else current_pid
                col_s, col_d = st.columns([1, 1])
                with col_s:
                    new_status = "対応済み" if status == "未対応" else "未対応"
                    if st.button(f"{new_status}にする", key=f"status_fb_{target_pid}_{idx}"):
                        proj_fbs = load_project_feedbacks(target_pid)
                        # タイムスタンプと内容で一致するものを探す
                        for i, pfb in enumerate(proj_fbs):
                            if pfb.get('timestamp') == fb.get('timestamp') and pfb.get('body') == fb.get('body'):
                                proj_fbs[i]['status'] = new_status
                                break
                        save_project_feedbacks(target_pid, proj_fbs)
                        st.rerun()
                with col_d:
                    if st.button("削除", key=f"del_fb_{target_pid}_{idx}", type="secondary"):
                        proj_fbs = load_project_feedbacks(target_pid)
                        for i, pfb in enumerate(proj_fbs):
                            if pfb.get('timestamp') == fb.get('timestamp') and pfb.get('body') == fb.get('body'):
                                proj_fbs.pop(i)
                                break
                        save_project_feedbacks(target_pid, proj_fbs)
                        st.rerun()

            st.markdown("")
            col1, col2 = st.columns([1, 3])
            with col1:
                if fb_view_mode != "全案件横断表示" and current_pid:
                    if st.button("この案件のFBをクリア", type="secondary"):
                        save_project_feedbacks(current_pid, [])
                        st.rerun()
            with col2:
                csv_fbs = display_fbs
                fb_csv_data = "日時,案件名,カテゴリ,対象,内容,優先度,送信者,物件名,ステータス\n"
                for fb in csv_fbs:
                    body_escaped = fb['body'].replace('"', '""').replace('\n', ' ')
                    target_escaped = fb.get('target', '').replace('"', '""')
                    pname = fb.get('_proj_name', current_pname).replace('"', '""')
                    fb_csv_data += (
                        f'"{fb["timestamp"]}","{pname}","{fb["category"]}","{target_escaped}",'
                        f'"{body_escaped}","{fb["priority"]}","{fb.get("name", "")}",'
                        f'"{fb.get("property", "")}","{fb.get("status", "未対応")}"\n'
                    )
                st.download_button(
                    label="フィードバック一覧をCSVダウンロード",
                    data=fb_csv_data.encode('utf-8-sig'),
                    file_name="feedback.csv",
                    mime="text/csv",
                    use_container_width=True,
                )
        else:
            if current_pid:
                st.info("この案件にはまだフィードバックがありません。")
            else:
                st.info("案件が選択されていません。CADデータをアップロードして案件を作成してください。")


if __name__ == "__main__":
    main()
